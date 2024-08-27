import openpyxl

from openpyxl.styles import Font    
from fastapi.responses import FileResponse
import datetime
from fastapi import HTTPException

from ...models.users.users import Users

from ...models.users.attached_branches import attached_branches
from ..loan_monitoring.report_order import report_order_crud
from ...common.dictionaries.general_tasks_dictionary import  MGT
from ...common.dictionaries.departments_dictionary import ROLES
from ...common.commit import commit_object, flush_object
from ...config.logs_config import info_logger
import uuid, json
from openpyxl.styles import Alignment

import time
from sqlalchemy.sql import text


def create_report_to_excel(request, user_id, db_session):
    response = check_user_acces_to_locals(user_id,request.report_type, db_session)
    report_path = f"project_files/loan_reports/loan_report_{str(datetime.datetime.now().date())}_{str(uuid.uuid4())[:4]}.xlsx"
    report_order_crud.report_order(user_id, report_path, request.report_type, request.report_by, db_session)
    period_name = 'loan_report'
    
    create_excel(report_path, user_id, response, db_session)
    report_order_crud.change_report_status_to_ready(user_id, request.report_type, request.report_by, db_session)
    commit_object(db_session)
    print('report ready')
    info_logger.info("User %s requested report for loan monitoring", user_id)
    return FileResponse(report_path, filename=f"{period_name+'_'+str(datetime.datetime.now().date())}.xlsx")
    




def create_excel(report_path, user_id, response, db_session):
    book=openpyxl.Workbook()
    sheet = book.active
    start_timer = time.time()
    locals_script=''
    from_clause_kad_case=''
    if response['is_main_superviser']:
        locals_script=f"AND LOAN_PORTFOLIO.LOCAL_CODE_ID {response['locals']}"
    else:
        from_clause_kad_case = f'''AND (LOAN_CASE.MAIN_RESPONSIBLE_ID={user_id} OR LOAN_CASE.SECOND_RESPONSIBLE_ID={user_id}) '''
    get_by_copy = db_session.execute(text(f'''SELECT
                                                LOAN_CASE.ID LOAN_CASE1_ID,
                                                LOAN_PORTFOLIO.CLIENT_NAME,
                                                LOAN_PORTFOLIO.LOAN_ID,
                                                LOAN_PORTFOLIO.LOAN_ACCOUNT,
                                                CLIENT_REGION.NAME AS CLIENT_REGION,
                                                LOAN_PORTFOLIO.ISSUE_DATE,
                                                LOAN_PORTFOLIO.MATURITY_DATE,
                                                (CASE WHEN LOAN_PORTFOLIO.CREDIT_ACCOUNT_BALANCE IS NOT NULL AND LOAN_PORTFOLIO.CREDIT_ACCOUNT_BALANCE!='' THEN LOAN_PORTFOLIO.CREDIT_ACCOUNT_BALANCE::real ELSE 0 END +
                                                CASE WHEN LOAN_PORTFOLIO.OVERDUE_BALANCE IS NOT NULL AND LOAN_PORTFOLIO.OVERDUE_BALANCE!='' THEN LOAN_PORTFOLIO.OVERDUE_BALANCE::real ELSE 0 END) ACCOUNT_BALANCE,
                                                LOAN_PORTFOLIO.LENDING_TYPE,
                                                LOCAL_CODE.CODE LOCAL_CODE,
                                                LOCAL_CODE.NAME LOCAL_CODE_NAME,
                                                MAIN_RESPONSIBLE.FULL_NAME MAIN_RESPONSIBLE,
                                                SECOND_RESPONSIBLE.FULL_NAME SECOND_RESPONSIBLE,
                                                
                                                TARGET_MONITORING.MAIN_RESPONSIBLE_DUE_DATE::DATE TARGET_MAIN_RESPONSIBLE_DUE_DATE,
                                                TARGET_MONITORING.SECOND_RESPONSIBLE_DUE_DATE::DATE TARGET_SECOND_RESPONSIBLE_DUE_DATE,
                                                TARGET_MONITORING.DEADLINE::DATE TARGET_DEADLINE,
                                                TARGET_MONITORING_STATUS.NAME TARGET_STATUS,
                                                TARGET_MONITORING.TARGET_MONITORING_RESULT_REASON_ID TARGET_MONITORING_RESULT_REASON,
                                                TARGET_RESULT.NAME AS TARGET_RESULT_NAME,
                                                TARGET_REASON.NAME AS TARGET_REASON_NAME,
                                                
                                                SCHEDULED_MONITORING.MAIN_RESPONSIBLE_DUE_DATE::DATE SCHEDULED_MAIN_RESPONSIBLE_DUE_DATE,
                                                SCHEDULED_MONITORING.SECOND_RESPONSIBLE_DUE_DATE::DATE SCHEDULED_SECOND_RESPONSIBLE_DUE_DATE,
                                                SCHEDULED_MONITORING.DEADLINE::DATE SCHEDULED_DEADLINE,
                                                SCHEDULED_MONITORING_STATUS.NAME SCHEDUELD_STATUS,
                                                SCHEDULED_MONITORING.SCHEDULED_MONITORING_RESULT_REASON_ID SCHEDULED_MONITORING_RESULT_REASON,
                                                SCHEDULED_RESULT.NAME AS SCHEDULED_RESULT_NAME,
                                                SCHEDULED_REASON.NAME AS SCHEDULED_REASON_NAME,
                                                
                                                UNSCHEDULED_MONITORING.MAIN_RESPONSIBLE_DUE_DATE::DATE UNSCHEDULED_MAIN_RESPONSIBLE_DUE_DATE,
                                                UNSCHEDULED_MONITORING.SECOND_RESPONSIBLE_DUE_DATE::DATE UNSCHEDULED_SECOND_RESPONSIBLE_DUE_DATE,
                                                UNSCHEDULED_MONITORING.DEADLINE::DATE UNSCHEDULED_DEADLINE,
                                                UNSCHEDULED_MONITORING_STATUS.NAME UNSCHEDUELD_STATUS,
                                                UNSCHEDULED_MONITORING.UNSCHEDULED_MONITORING_RESULT_REASON_ID UNSCHEDULED_MONITORING_RESULT_REASON,
                                                UNSCHEDULED_RESULT.NAME AS UNSCHEDULED_RESULT_NAME,
                                                UNSCHEDULED_REASON.NAME AS UNSCHEDULED_REASON_NAME,
                                                
                                                
                                                LOAN_PORTFOLIO.CONTRACT_AMOUNT,
                                                LOAN_PORTFOLIO.OSN_CMP_PERCENT,
                                                CURRENCY.NAME CURRENCY,
                                                CURRENCY.CODE CURRENCY_CODE,
                                                MONITORING_CASE.MONITORING_CASE_STATUS_ID
                                                
                                            FROM LOAN_CASE LOAN_CASE
                                            JOIN LOAN_PORTFOLIO LOAN_PORTFOLIO ON LOAN_PORTFOLIO.ID = LOAN_CASE.LOAN_PORTFOLIO_ID
                                            JOIN CLIENT_REGION CLIENT_REGION ON LOAN_PORTFOLIO.CLIENT_REGION = CLIENT_REGION.ID
                                            JOIN LOCAL_CODE LOCAL_CODE ON LOCAL_CODE.ID = LOAN_PORTFOLIO.LOCAL_CODE_ID
                                            JOIN CURRENCY CURRENCY ON LOAN_PORTFOLIO.CURRENCY_ID = CURRENCY.ID
                                            LEFT JOIN MONITORING_CASE MONITORING_CASE ON MONITORING_CASE.ID = LOAN_CASE.MONITORING_CASE_ID
                                            JOIN USERS MAIN_RESPONSIBLE ON MAIN_RESPONSIBLE.ID = LOAN_CASE.MAIN_RESPONSIBLE_ID
                                            JOIN USERS SECOND_RESPONSIBLE ON SECOND_RESPONSIBLE.ID = LOAN_CASE.SECOND_RESPONSIBLE_ID
                                            LEFT JOIN TARGET_MONITORING TARGET_MONITORING ON TARGET_MONITORING.ID = MONITORING_CASE.TARGET_MONITORING_ID
                                            LEFT JOIN MONITORING_STATUS TARGET_MONITORING_STATUS ON TARGET_MONITORING_STATUS.ID = TARGET_MONITORING.TARGET_MONITORING_STATUS_ID
                                            LEFT JOIN TARGET_MONITORING_RESULT TARGET_RESULT ON TARGET_RESULT.ID = TARGET_MONITORING.TARGET_MONITORING_RESULT_ID
                                            
                                            LEFT JOIN RESULT_REASON TARGET_REASON ON TARGET_REASON.ID = TARGET_MONITORING.TARGET_MONITORING_RESULT_REASON_ID
                                            
                                            LEFT JOIN
                                                (SELECT *
                                                    FROM SCHEDULED_MONITORING SCHEDULED_MONITORING_1
                                                    WHERE ID =
                                                            (SELECT MAX(ID)
                                                                FROM SCHEDULED_MONITORING SCHEDULED_MONITORING_2
                                                                WHERE SCHEDULED_MONITORING_1.MONITORING_CASE_ID = SCHEDULED_MONITORING_2.MONITORING_CASE_ID)) SCHEDULED_MONITORING ON SCHEDULED_MONITORING.MONITORING_CASE_ID = LOAN_CASE.MONITORING_CASE_ID
                                            LEFT JOIN MONITORING_STATUS SCHEDULED_MONITORING_STATUS ON SCHEDULED_MONITORING_STATUS.ID = SCHEDULED_MONITORING.SCHEDULED_MONITORING_STATUS_ID


                                            LEFT JOIN
                                                (SELECT *
                                                    FROM UNSCHEDULED_MONITORING UNSCHEDULED_MONITORING_1
                                                    WHERE ID =
                                                            (SELECT MAX(ID)
                                                                FROM UNSCHEDULED_MONITORING UNSCHEDULED_MONITORING_2
                                                                WHERE UNSCHEDULED_MONITORING_1.MONITORING_CASE_ID = UNSCHEDULED_MONITORING_2.MONITORING_CASE_ID)) UNSCHEDULED_MONITORING ON UNSCHEDULED_MONITORING.MONITORING_CASE_ID = LOAN_CASE.MONITORING_CASE_ID
                                            LEFT JOIN MONITORING_STATUS UNSCHEDULED_MONITORING_STATUS ON UNSCHEDULED_MONITORING_STATUS.ID = UNSCHEDULED_MONITORING.UNSCHEDULED_MONITORING_STATUS_ID
                                            LEFT JOIN TARGET_MONITORING_RESULT SCHEDULED_RESULT ON SCHEDULED_RESULT.ID = SCHEDULED_MONITORING.SCHEDULED_MONITORING_RESULT_ID
                                            LEFT JOIN TARGET_MONITORING_RESULT UNSCHEDULED_RESULT ON UNSCHEDULED_RESULT.ID = UNSCHEDULED_MONITORING.UNSCHEDULED_MONITORING_RESULT_ID
                                            LEFT JOIN RESULT_REASON SCHEDULED_REASON ON SCHEDULED_REASON.ID = SCHEDULED_MONITORING.SCHEDULED_MONITORING_RESULT_REASON_ID
                                            LEFT JOIN RESULT_REASON UNSCHEDULED_REASON ON UNSCHEDULED_REASON.ID = UNSCHEDULED_MONITORING.UNSCHEDULED_MONITORING_RESULT_REASON_ID
                                            WHERE LOAN_PORTFOLIO.STATUS=1 AND LOAN_PORTFOLIO.IS_TAKEN_LOAN=TRUE AND LOAN_PORTFOLIO.TOTAL_OVERDUE IS NOT NULL AND MONITORING_CASE.MONITORING_CASE_STATUS_ID=1
                                            {locals_script}
                                            {from_clause_kad_case}
                                          ''')).fetchall()

 
    print(time.time() - start_timer)
    row = 3
    i=1
    for loan in get_by_copy:
        if loan.monitoring_case_status_id == 2:
            target = {"deadline": None,
            "monitoring_status": 'Мониторинг не требуется',
            "task_status":None,
            "main_responsible_due_date": None,
            "main_responsible_expired": None,
            "second_responsible_due_date": None,
            "second_responsible_expired": None,
            "expired": None}
        else:
            target = get_target_monitoring_deadline(*loan[13:17])
        schedule = get_plan_monitoring_deadline(*loan[20:24])
        unschedule = get_unscheduled_monitoring_deadline(*loan[27:31])

        
        sheet.cell(column=1, row=row).value = i
        sheet.cell(column=2, row=row).value = loan['client_name']
        sheet.cell(column=3, row=row).value = loan['loan_id']
        sheet.cell(column=4, row=row).value = loan['loan_account']
        sheet.cell(column=5, row=row).value = loan['client_region']
        sheet.cell(column=6, row=row).value = loan['local_code_name']
        sheet.cell(column=7, row=row).value = loan['local_code']
        sheet.cell(column=8, row=row).value = '{0:,}'.format(float(loan['contract_amount'])).replace(',', '')
        sheet.cell(column=9, row=row).value = loan['issue_date']
        sheet.cell(column=10, row=row).value = loan['maturity_date']
        sheet.cell(column=11, row=row).value = '{0:,}'.format(float(loan['account_balance']!=None and loan['account_balance'] or 0)).replace(',', '')
        sheet.cell(column=12, row=row).value = loan['currency_code']
        sheet.cell(column=13, row=row).value = str(loan['osn_cmp_percent']) + '%'
        sheet.cell(column=14, row=row).value = loan['lending_type']
        
        #target
        sheet.cell(column=15, row=row).value = loan['main_responsible']
        sheet.cell(column=16, row=row).value = target['monitoring_status']
        sheet.cell(column=17, row=row).value = loan['target_deadline']
        sheet.cell(column=18, row=row).value = loan['target_second_responsible_due_date']
        sheet.cell(column=19, row=row).value = target['second_responsible_expired']
        sheet.cell(column=20, row=row).value = loan['second_responsible']
        sheet.cell(column=21, row=row).value = loan['target_main_responsible_due_date']
        sheet.cell(column=22, row=row).value = target['main_responsible_expired']
        sheet.cell(column=23, row=row).value = target['task_status']
        sheet.cell(column=24, row=row).value = loan['target_result_name']
        sheet.cell(column=25, row=row).value = loan['target_reason_name']
        
        #schedule
        sheet.cell(column=26, row=row).value = schedule['monitoring_status']
        sheet.cell(column=27, row=row).value = loan['scheduled_deadline']
        sheet.cell(column=28, row=row).value = loan['scheduled_second_responsible_due_date']
        sheet.cell(column=29, row=row).value = schedule['second_responsible_expired']
        sheet.cell(column=30, row=row).value = loan['second_responsible']
        sheet.cell(column=31, row=row).value = loan['scheduled_main_responsible_due_date']
        sheet.cell(column=32, row=row).value = schedule['main_responsible_expired']
        sheet.cell(column=33, row=row).value = schedule['task_status']
        sheet.cell(column=34, row=row).value = loan['scheduled_result_name']
        sheet.cell(column=35, row=row).value = loan['scheduled_reason_name']
        
        
        #unschedule
        sheet.cell(column=36, row=row).value = unschedule['monitoring_status']
        sheet.cell(column=37, row=row).value = loan['unscheduled_deadline']
        sheet.cell(column=38, row=row).value = loan['unscheduled_second_responsible_due_date']
        sheet.cell(column=39, row=row).value = unschedule['second_responsible_expired']
        sheet.cell(column=40, row=row).value = loan['second_responsible']
        sheet.cell(column=41, row=row).value = loan['unscheduled_main_responsible_due_date']
        sheet.cell(column=42, row=row).value = unschedule['main_responsible_expired']
        sheet.cell(column=43, row=row).value = unschedule['task_status']
        sheet.cell(column=44, row=row).value = loan['unscheduled_result_name']
        sheet.cell(column=45, row=row).value = loan['unscheduled_reason_name']
        
        row = row + 1
        i = i + 1
    
    #sheet.insert_rows(1)
    
    sheet.cell(column=1, row=1).value = "№"
    sheet.cell(column=2, row=1).value = "Наименование клиента"
    sheet.cell(column=3, row=1).value = "ID договор"
    sheet.cell(column=4, row=1).value = "Кредитный счёт"
    sheet.cell(column=5, row=1).value = "Регион"
    sheet.cell(column=6, row=1).value = "Наименование ЦБУ, ОФБ"
    sheet.cell(column=7, row=1).value = "Локал код"
    sheet.cell(column=8, row=1).value = "Сумма договора"
    sheet.cell(column=9, row=1).value = "Дата выдачи"
    sheet.cell(column=10, row=1).value = "Дата погащения"
    sheet.cell(column=11, row=1).value = "Остаток кредитного счёта"
    sheet.cell(column=12, row=1).value = "Валюта"
    sheet.cell(column=13, row=1).value = "OSN+CMP Процент кредита"
    sheet.cell(column=14, row=1).value = "Вид кредитования"
    
    
    
    sheet.cell(column=16, row=1).value = "Целевой мониторинг"
    ##################################################################
    sheet.cell(column=15, row=1).value = "Ответственный сотрудник (Головной офис)"
    sheet.cell(column=16, row=2).value = "Статус мониторинга"
    sheet.cell(column=17, row=2).value = "Срок выполнения мониторинга (ЦБУ, ОФБ)"
    sheet.cell(column=18, row=2).value = "Дата проведения мониторинга (ЦБУ, ОФБ)"
    sheet.cell(column=19, row=2).value = "Просроченные дни (ЦБУ, ОФБ)"
    sheet.cell(column=20, row=2).value = "Ответственный сотрудник  (ЦБУ, ОФБ)"
    sheet.cell(column=21, row=2).value = "Дата проверки (Головной офис)"
    sheet.cell(column=22, row=2).value = "Просроченные дни (Головной офис)"
    sheet.cell(column=23, row=2).value = "Статус задачи"
    sheet.cell(column=24, row=2).value = "Статус мониторинга (Целевой, частично целевые, не целевой)"
    sheet.cell(column=25, row=2).value = "Состояние кредита по результатам мониторинга"
    
    
    sheet.cell(column=26, row=1).value = "Плановый мониторинг"
    #################################################################
    sheet.cell(column=26, row=2).value = "Статус мониторинга"
    sheet.cell(column=27, row=2).value = "Срок выполнения мониторинга (ЦБУ, ОФБ)"
    sheet.cell(column=28, row=2).value = "Дата проведения мониторинга (ЦБУ, ОФБ)"
    sheet.cell(column=29, row=2).value = "Просроченные дни (ЦБУ, ОФБ)"
    sheet.cell(column=30, row=2).value = "Ответственный сотрудник  (ЦБУ, ОФБ)"
    sheet.cell(column=31, row=2).value = "Дата проверки (Головной офис)"
    sheet.cell(column=32, row=2).value = "Просроченные дни (Головной офис)"
    sheet.cell(column=33, row=2).value = "Статус задачи"
    sheet.cell(column=34, row=2).value = "Статус мониторинга (Целевой, частично целевые, не целевой)"
    sheet.cell(column=35, row=2).value = "Состояние кредита по результатам мониторинга"
    
    
    sheet.cell(column=36, row=1).value = "Внеплановый мониторинг"
    ##################################################################
    sheet.cell(column=36, row=2).value = "Статус мониторинга"
    sheet.cell(column=37, row=2).value = "Срок выполнения мониторинга (ЦБУ, ОФБ)"
    sheet.cell(column=38, row=2).value = "Дата проведения мониторинга (ЦБУ, ОФБ)"
    sheet.cell(column=39, row=2).value = "Просроченные дни (ЦБУ, ОФБ)"
    sheet.cell(column=40, row=2).value = "Ответственный сотрудник  (ЦБУ, ОФБ)"
    sheet.cell(column=41, row=2).value = "Дата проверки (Головной офис)"
    sheet.cell(column=42, row=2).value = "Просроченные дни (Головной офис)"
    sheet.cell(column=43, row=2).value = "Статус задачи"
    sheet.cell(column=44, row=2).value = "Статус мониторинга (Целевой, частично целевые, не целевой)"
    sheet.cell(column=45, row=2).value = "Состояние кредита по результатам мониторинга"
   
    bold = Font(bold=True)

    # for cell in sheet["1:1"]:
    #     cell.font = bold
    # sheet.row_dimensions[1].height = 30
    # sheet.merge_cells('I1:M1')
    # megre_cell = sheet['I1']
    # megre_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    
    # sheet.merge_cells('N1:R1')
    # megre_cell = sheet['N1']
    # megre_cell.alignment = Alignment(horizontal="center", vertical="center")
    # sheet.merge_cells('S1:W1')
    # megre_cell = sheet['S1']
    # megre_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    
    for row in sheet["1:2"]:
        for cell in row:
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    sheet.merge_cells('A1:A2')
    sheet.merge_cells('B1:B2')
    sheet.merge_cells('C1:C2')
    sheet.merge_cells('D1:D2')
    sheet.merge_cells('E1:E2')
    sheet.merge_cells('F1:F2')
    sheet.merge_cells('G1:G2')
    sheet.merge_cells('H1:H2')
    sheet.merge_cells('I1:I2')
    sheet.merge_cells('J1:J2')
    sheet.merge_cells('K1:K2')
    sheet.merge_cells('L1:L2')
    sheet.merge_cells('M1:M2')
    sheet.merge_cells('N1:N2')
    sheet.merge_cells('O1:O2')
    sheet.merge_cells('P1:Y1')
    sheet.merge_cells('Z1:AI1')
    sheet.merge_cells('AJ1:AS1')
    
    sheet.column_dimensions["A"].width=8
    sheet.column_dimensions["B"].width=50
    sheet.column_dimensions["C"].width=13
    sheet.column_dimensions["D"].width=25
    sheet.column_dimensions["E"].width=30
    sheet.column_dimensions["F"].width=30
    sheet.column_dimensions["G"].width=10
    sheet.column_dimensions["H"].width=15
    sheet.column_dimensions["I"].width=12
    sheet.column_dimensions["J"].width=12
    sheet.column_dimensions["K"].width=15
    sheet.column_dimensions["L"].width=10
    sheet.column_dimensions["M"].width=13
    sheet.column_dimensions["N"].width=28
    sheet.column_dimensions["O"].width=25
    sheet.column_dimensions["P"].width=25
    sheet.column_dimensions["Q"].width=25
    sheet.column_dimensions["R"].width=37
    sheet.column_dimensions["S"].width=20
    sheet.column_dimensions["T"].width=20
    sheet.column_dimensions["U"].width=25
    sheet.column_dimensions["V"].width=25
    sheet.column_dimensions["W"].width=12
    sheet.column_dimensions["X"].width=37
    sheet.column_dimensions["Y"].width=35
    sheet.column_dimensions["Z"].width=25
    sheet.column_dimensions["AA"].width=25
    sheet.column_dimensions["AB"].width=25
    sheet.column_dimensions["AC"].width=17
    sheet.column_dimensions["AD"].width=37
    sheet.column_dimensions["AE"].width=20
    sheet.column_dimensions["AF"].width=20
    sheet.column_dimensions["AG"].width=13
    sheet.column_dimensions["AH"].width=37
    sheet.column_dimensions["AI"].width=23
    sheet.column_dimensions["AJ"].width=25
    sheet.column_dimensions["AK"].width=25
    sheet.column_dimensions["AL"].width=25
    sheet.column_dimensions["AM"].width=17
    sheet.column_dimensions["AN"].width=37
    sheet.column_dimensions["AO"].width=20
    sheet.column_dimensions["AP"].width=20
    sheet.column_dimensions["AQ"].width=13
    sheet.column_dimensions["AR"].width=37
    sheet.column_dimensions["AS"].width=23
    
    
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 40
    
    book.save(report_path)
    print('end: ', time.time() - start_timer, " sec")
    




def get_monitoring_stage(general_task_id, is_taken_loan, is_taken_problem, is_taken_juridic, is_taken_business):
    if (is_taken_problem and is_taken_juridic) or\
        (not is_taken_problem and is_taken_juridic) or\
            (not is_taken_juridic and general_task_id == MGT.SEND_JURIDIC.value):
        monitoring_stage = 'Юристы'
    elif (is_taken_problem and not is_taken_juridic == False) or\
        (is_taken_problem and not is_taken_loan == False) or\
            (not is_taken_problem and general_task_id == MGT.SEND_PORBLEM.value):
        monitoring_stage = 'Проблемные активы'
    else:
        monitoring_stage = 'Мониторинг'
    
    return monitoring_stage

    
def get_target_monitoring_deadline(main_responsible_due_date, second_responsible_due_date, deadline, task_status):
    monitoring_status = None
    expired = False
    main_responsible_expired = None
    second_responsible_expired = None
    if deadline is not None:
        monitoring_status = 'Целевой мониторинг'
        if second_responsible_due_date is not None:
            if (second_responsible_due_date - deadline).days > 0:
                monitoring_status = 'Целевой с просрочкой'
                expired = True
                second_responsible_expired = (second_responsible_due_date - deadline).days

            if main_responsible_due_date is not None:
                if (main_responsible_due_date - second_responsible_due_date).days > 2:
                    main_responsible_expired = (main_responsible_due_date - second_responsible_due_date).days - 2
         
    else:
        monitoring_status = 'Мониторинг не требуется'
    return {"deadline": deadline,
            "monitoring_status": monitoring_status,
            "task_status":task_status,
            "main_responsible_due_date": main_responsible_due_date,
            "main_responsible_expired": main_responsible_expired,
            "second_responsible_due_date": second_responsible_due_date,
            "second_responsible_expired": second_responsible_expired,
            "expired": expired}
    

def get_plan_monitoring_deadline(main_responsible_due_date, second_responsible_due_date, deadline, task_status):
    monitoring_status = ''
    main_responsible_expired = None
    second_responsible_expired = None
    expired = False
    if deadline is not None:
        monitoring_status = 'Плановый мониторинг'
        if second_responsible_due_date is not None:
            if (second_responsible_due_date - deadline).days > 0:
                monitoring_status = 'Плановый с просрочкой'
                expired = True
                second_responsible_expired = (second_responsible_due_date - deadline).days

            if main_responsible_due_date is not None:
                if (main_responsible_due_date - second_responsible_due_date).days > 2:
                    main_responsible_expired = (main_responsible_due_date - second_responsible_due_date).days - 2


    return {"deadline": deadline,
            "monitoring_status": monitoring_status,
            "task_status": task_status,
            "main_responsible_due_date": main_responsible_due_date,
            "main_responsible_expired": main_responsible_expired,
            "second_responsible_due_date": second_responsible_due_date,
            "second_responsible_expired": second_responsible_expired,
            "expired": expired}    
    
    
    
def get_unscheduled_monitoring_deadline(main_responsible_due_date, second_responsible_due_date, deadline, task_status):
    monitoring_status = ''
    main_responsible_expired = None
    second_responsible_expired = None
    expired = False
    if deadline is not None:
        monitoring_status = 'Внеплановый мониторинг'
        if second_responsible_due_date is not None:
            if (second_responsible_due_date - deadline).days > 0:
                monitoring_status = 'Внеплановый с просрочкой'
                expired = True
                second_responsible_expired = (second_responsible_due_date - deadline).days

            if main_responsible_due_date is not None:
                if (main_responsible_due_date - second_responsible_due_date).days > 2:
                    main_responsible_expired = (main_responsible_due_date - second_responsible_due_date).days - 2


    return {"deadline": deadline,
            "monitoring_status": monitoring_status,
            "task_status": task_status,
            "main_responsible_due_date": main_responsible_due_date,
            "main_responsible_expired": main_responsible_expired,
            "second_responsible_due_date": second_responsible_due_date,
            "second_responsible_expired": second_responsible_expired,
            "expired": expired} 
    
    

MAIN_ADMINS = [ROLES.monitoring_main_admin.value, ROLES.business_block_main_admin.value,
              ROLES.kad_block_main_admin.value, ROLES.promlem_block_main_admin.value]    

FILIAL_USERS = [ROLES.monitoring_filial_admin.value, ROLES.monitoring_filial_user.value,
                ROLES.kad_block_filial_admin.value, ROLES.kad_block_filial_user.value,
                ROLES.business_block_filial_admin.value, ROLES.business_block_filial_user.value,
                ROLES.problem_block_filial_admin.value, ROLES.promlem_block_filial_user.value]

MAIN_SUPERVISERS = [ROLES.main_superviser_kad.value, ROLES.main_superviser_business.value,
                    ROLES.main_superviser_problem.value,ROLES.main_superviser_kad_with_pledge.value,
                    ROLES.main_superviser.value, ROLES.superviser.value]






    
def check_user_acces_to_locals(user_id, report_type, db_session):
    get_user = db_session.query(Users).filter(Users.id==user_id).first()
    user_roles = []
    
    for role in get_user.roles:
        user_roles.append(role.id)
    
    fl_main_admin = False
    for role in user_roles:
        if role in MAIN_ADMINS:
            fl_main_admin = True
    
    fl_filial_users = False
    for role in user_roles:
        if role in FILIAL_USERS:
            fl_filial_users = True
    
    
    fl_main_supervisers = False
    for role in user_roles:
        if role in MAIN_SUPERVISERS:
            fl_main_supervisers = True
    
    
    if fl_filial_users:
        return {"locals":'='+ str(get_user.local_code), "is_main_superviser":fl_main_supervisers}
    
    if fl_main_admin:
        locals = db_session.query(attached_branches.local_code_id).filter(attached_branches.user_id==get_user.id).filter(attached_branches.department_id==get_user.department)
    elif fl_main_supervisers:
        locals = db_session.query(attached_branches.local_code_id).filter(attached_branches.department_id==get_user.department)
    if report_type == 1:
        locals = locals.filter(attached_branches.attached_type_id==3).all()
    elif report_type == 2:
        locals = locals.filter(attached_branches.attached_type_id==1).all()
    elif report_type == 3:
        locals = locals.filter(attached_branches.attached_type_id==2).all()
    elif report_type == 4 or report_type == 5:
        locals = locals.filter(attached_branches.attached_type_id==4).all()
        
    if len(tuple(locals))==1:
        return {"locals":'='+ str(get_user.local_code), 
                "is_main_superviser":fl_main_supervisers}
    else:
        return {"locals":'IN '+ str(tuple(item[0] for item in locals)),"is_main_superviser":fl_main_supervisers}
    
    
    