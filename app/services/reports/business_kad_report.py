import openpyxl

from openpyxl.styles import Font, Alignment
from fastapi.responses import FileResponse
import datetime
from fastapi import HTTPException
from .report_to_excel_service import check_user_acces_to_locals
from ..loan_monitoring.report_order import report_order_crud
from ...models.statuses.from_type_model import FromType
from ...common.dictionaries.general_tasks_dictionary import  MGT
from ...common.commit import commit_object, flush_object
from ...common.decorator import measure_time
from ...config.logs_config import info_logger
import uuid, json
import time
from sqlalchemy.sql import text
user_id = None
BUSINESS_REPORT_TYPE = f'''JOIN BUSINESS_CASE BC ON LP.ID = BC.LOAN_PORTFOLIO_ID 
                            JOIN FROM_TYPE FT ON BC.FROM_TYPE_ID = FT.ID
                            WHERE LP.IS_TAKEN_BUSINESS = true
                            '''

KAD_REPORT_TYPE = f'''JOIN KAD_CASE KC ON LP.ID = KC.LOAN_PORTFOLIO_ID 
                        JOIN FROM_TYPE FT ON KC.FROM_TYPE_ID = FT.ID
                        LEFT JOIN HYBRID_LETTERS HL1 ON HL1.KAD_CASE_ID = KC.ID
                        LEFT JOIN HYBRID_LETTERS HL2 ON HL2.KAD_CASE_ID = KC.ID
                        LEFT JOIN LETTER_STATUS LS1 ON LS1.ID = HL1.LETTER_STATUS_ID
                        LEFT JOIN LETTER_STATUS LS2 ON LS2.ID = HL2.LETTER_STATUS_ID
                            WHERE LP.IS_TAKEN_KAD = true
                            '''

PROBLEM_REPORT_TYPE = f'''JOIN PROBLEMS_CASE PC ON LP.ID = PC.LOAN_PORTFOLIO_ID 
                        JOIN FROM_TYPE FT ON PC.FROM_TYPE_ID = FT.ID
                            WHERE LP.IS_TAKEN_PROBLEM = true AND LP.IS_TAKEN_OUT_OF_BALANCE = false
                            '''

OUT_OF_BALANCE_REPORT_TYPE = f'''JOIN PROBLEMS_CASE PC ON LP.ID = PC.LOAN_PORTFOLIO_ID 
                        JOIN FROM_TYPE FT ON PC.FROM_TYPE_ID = FT.ID
                            WHERE LP.IS_TAKEN_PROBLEM = true AND LP.IS_TAKEN_OUT_OF_BALANCE = true
                            '''

@measure_time
def create_report_to_excel(request, user_id, db_session):
    response = check_user_acces_to_locals(user_id, request.report_type, db_session)
    
    report_type=''
    if request.report_type == 2:
        if response['is_main_superviser'] is False:
            report_type = BUSINESS_REPORT_TYPE +  f'AND (BC.MAIN_RESPONSIBLE_ID = {user_id} OR BC.SECOND_RESPONSIBLE_ID = {user_id})'
        else:
            report_type = BUSINESS_REPORT_TYPE
        report_name = 'business'
        
    elif request.report_type == 3:
        if response['is_main_superviser'] is False:
            report_type = KAD_REPORT_TYPE + f'AND (KC.MAIN_RESPONSIBLE_ID = {user_id} OR KC.SECOND_RESPONSIBLE_ID = {user_id})'
        else:
            report_type = KAD_REPORT_TYPE
        report_name = 'kad'
        
    elif request.report_type == 4:
        if response['is_main_superviser'] is False:
            report_type = PROBLEM_REPORT_TYPE + f'AND (PC.MAIN_RESPONSIBLE_ID = {user_id} OR PC.SECOND_RESPONSIBLE_ID = {user_id})'
        else:
            report_type = PROBLEM_REPORT_TYPE
        report_name = 'problems'
        
    elif request.report_type == 5:
        if response['is_main_superviser'] is False:
            report_type = OUT_OF_BALANCE_REPORT_TYPE + f'AND (PC.MAIN_RESPONSIBLE_ID = {user_id} OR PC.SECOND_RESPONSIBLE_ID = {user_id})'
        else:
            report_type = OUT_OF_BALANCE_REPORT_TYPE
        report_name = 'out_of_bal'
    
    report_path = f"project_files/loan_reports/{report_name}_{str(datetime.datetime.now().date())}_{str(uuid.uuid4())[:4]}.xlsx"
    report_order_crud.report_order(user_id, report_path, request.report_type, request.report_by, db_session)
    
    
        
    if request.report_by == 1:
        report_by = '_by_client'
        if request.report_type == 2:
            create_excel_by_client_business(report_path, report_type, response['locals'], db_session)
        elif request.report_type == 3:
            create_excel_by_client_kad(report_path, report_type, response['locals'], db_session)
        elif request.report_type == 4:
            create_excel_by_client_problems(report_path, report_type, response['locals'], db_session)
        elif request.report_type == 5:
            create_excel_by_client_out_of_balance(report_path, report_type, response['locals'], db_session)
    else:
        report_by = '_by_local'
        if request.report_type == 2:
            create_excel_by_local_business(report_path, report_type, response['locals'], db_session)
        if request.report_type == 3:
            create_excel_by_local_kad(report_path, report_type, response['locals'], db_session)
        if request.report_type == 4:
            create_excel_by_local_problems(report_path, report_type, response['locals'], db_session)
        if request.report_type == 5:
            create_excel_by_local_out_of_balance(report_path, report_type, response['locals'], db_session)
        
    report_order_crud.change_report_status_to_ready(user_id, request.report_type, request.report_by, db_session)
    commit_object(db_session)
    info_logger.info("User %s requested report for business or kad or problems", user_id)
    return FileResponse(report_path, filename=f"{report_name+report_by+'_'+str(datetime.datetime.now().date())}.xlsx")
    



def create_excel_by_client_business(report_path, report_type, locals, db_session):
    book=openpyxl.Workbook()
    sheet = book.active
    get_by_copy = db_session.execute(text(f''' 
                                          SELECT BT.ID,
                                            CR.NAME AS REGION,
                                            LC.CODE AS LOCAL_CODE,
                                            LP.LOAN_ID,
                                            LP.CLIENT_NAME,
                                            FT.ID AS FROM_TYPE_ID,
                                            BT.DEBT_ACCOUNT_START_STATE,
                                            (CASE WHEN BT.DEBT_ACCOUNT_CREDIT_SUM IS NOT NULL THEN (BT.DEBT_ACCOUNT_CREDIT_SUM::real)/100 ELSE 0 END) as DEBT_ACCOUNT_CREDIT_SUM,
                                            (
                                            CASE WHEN BT.DEBT_ACCOUNT_START_STATE IS NOT NULL THEN BT.DEBT_ACCOUNT_START_STATE::real ELSE 0 END
                                            -
                                            CASE WHEN BT.DEBT_ACCOUNT_CREDIT_SUM IS NOT NULL THEN (BT.DEBT_ACCOUNT_CREDIT_SUM::real)/100 ELSE 0 END) as DEBT_ACCOUNT_BALANCE,
                                            BT.ACCOUNT_16377_START_STATE,
                                            (CASE WHEN BT.ACCOUNT_16377_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_DEBIT_SUM,
                                            (CASE WHEN BT.ACCOUNT_16377_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_CREDIT_SUM,
                                            ((
                                            CASE WHEN BT.ACCOUNT_16377_START_STATE IS NOT NULL THEN BT.ACCOUNT_16377_START_STATE::real ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_16377_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_16377_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_BALANCE,
                                            
                                            (CASE WHEN BT.ACCOUNT_163XX_START_STATE IS NOT NULL THEN (BT.ACCOUNT_163XX_START_STATE::real) ELSE 0 END) as ACCOUNT_163XX_START_STATE,
                                            (CASE WHEN BT.ACCOUNT_163XX_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_DEBIT_SUM,
                                            (CASE WHEN BT.ACCOUNT_163XX_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_CREDIT_SUM,
                                            
                                            ((CASE WHEN BT.ACCOUNT_163XX_START_STATE IS NOT NULL THEN (BT.ACCOUNT_163XX_START_STATE::real) ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_163XX_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_163XX_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_BALANCE
                                        FROM BALANCE_TURNOVER BT
                                        LEFT JOIN LOAN_PORTFOLIO LP ON BT.LOAN_ID = LP.LOAN_ID
                                        LEFT JOIN CLIENT_REGION CR ON LP.CLIENT_REGION = CR.ID
                                        LEFT JOIN LOCAL_CODE LC ON LP.LOCAL_CODE_ID = LC.ID
                                        {report_type}
                                        AND LP.STATUS=1
                                        AND LP.LOCAL_CODE_ID {locals}
                                        
                                          ''')).fetchall()

    
    row = 3
    i=1
    
    for loan in get_by_copy:
        i=i+1
        
        sheet.cell(column=1, row=row).value = i
        sheet.cell(column=2, row=row).value = loan['region']
        sheet.cell(column=3, row=row).value = loan['local_code']
        sheet.cell(column=4, row=row).value = loan['loan_id']
        sheet.cell(column=5, row=row).value = loan['client_name']
        sheet.cell(column=6, row=row).value = loan['debt_account_start_state']
        if loan['from_type_id'] == 1:
            sheet.cell(column=7, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 2:
            sheet.cell(column=8, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 3:
            sheet.cell(column=9, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 4:
            sheet.cell(column=10, row=row).value = loan['debt_account_start_state']
        sheet.cell(column=11, row=row).value = loan['debt_account_credit_sum']
        sheet.cell(column=12, row=row).value = loan['debt_account_balance']
        sheet.cell(column=13, row=row).value = loan['account_16377_start_state']
        sheet.cell(column=14, row=row).value = loan['account_16377_debit_sum']
        sheet.cell(column=15, row=row).value = loan['account_16377_credit_sum']
        sheet.cell(column=16, row=row).value = loan['account_16377_balance']
        sheet.cell(column=17, row=row).value = loan['account_163xx_start_state']
        sheet.cell(column=18, row=row).value = loan['account_163xx_debit_sum']
        sheet.cell(column=19, row=row).value = loan['account_163xx_credit_sum']
        sheet.cell(column=20, row=row).value = loan['account_163xx_balance']
        
        row = row + 1
    
    #sheet.insert_rows(1)
    sheet.cell(column=7, row=1).value = "Остаток кредитов с просрочкой до 30 дней"
    sheet.cell(column=8, row=1).value = "Остаток кредитов с просрочкой до 30 дней из отчета за прошлый месяц"
    sheet.cell(column=9, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 31-60 дней"
    sheet.cell(column=10, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 60+  дней"
    
    sheet.insert_rows(1)
    sheet.cell(column=6, row=1).value = "Остаток кредитов к началу отчетного периода"
    sheet.cell(column=7, row=1).value = "В том числе"
    sheet.cell(column=11, row=1).value = "Погашено кредитов за отчетный период"
    sheet.cell(column=12, row=1).value = "Остаток кредитов к концу отчетного периода"
    
    sheet.insert_rows(1)
    sheet.cell(column=1, row=1).value = "ID"
    sheet.cell(column=2, row=1).value =  "Регион"
    sheet.cell(column=3, row=1).value = "Локал код"
    sheet.cell(column=4, row=1).value = "ID кредита"
    sheet.cell(column=5, row=1).value = "Наименование клиента"
    sheet.cell(column=6, row=1).value = "Задолженность с просрочкой до 30 дней (0-30)"
    sheet.cell(column=13, row=1).value = "Остаток 16377 на начало отчетного периода"
    sheet.cell(column=14, row=1).value = "Зачислено в течении месяца по счёту 16377"
    sheet.cell(column=15, row=1).value = "Погашено за отчетный период по счёту 16377"
    sheet.cell(column=16, row=1).value = "Остаток 16377 к концу отчетного периода"
    sheet.cell(column=17, row=1).value = "Остаток 163% на начало отчетного периода (кроме 16377)"
    sheet.cell(column=18, row=1).value = "Зачислено в течении месяца по счетам 163% (кроме 16377)"
    sheet.cell(column=19, row=1).value = "Погашено за отчетный период по счетам 163% (кроме 16377)"
    sheet.cell(column=20, row=1).value = "Остаток 163% к концу отчетного периода (кроме 16377)"
    
    bold = Font(bold=True)

    for row in sheet["1:3"]:
        for cell in row:
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    sheet.merge_cells('G2:J2')
    sheet.merge_cells('F1:L1')
    sheet.merge_cells('A1:A3')
    sheet.merge_cells('B1:B3')
    sheet.merge_cells('C1:C3')
    sheet.merge_cells('D1:D3')
    sheet.merge_cells('E1:E3')
    sheet.merge_cells('F2:F3')
    sheet.merge_cells('K2:K3')
    sheet.merge_cells('L2:L3')
    sheet.merge_cells('M1:M3')
    sheet.merge_cells('N1:N3')
    sheet.merge_cells('O1:O3')
    sheet.merge_cells('P1:P3')
    sheet.merge_cells('Q1:Q3')
    sheet.merge_cells('R1:R3')
    sheet.merge_cells('S1:S3')
    sheet.merge_cells('T1:T3')
    
    sheet.column_dimensions["A"].width=8
    sheet.column_dimensions["B"].width=8
    sheet.column_dimensions["C"].width=8
    sheet.column_dimensions["D"].width=11
    sheet.column_dimensions["E"].width=33
    sheet.column_dimensions["F"].width=20
    sheet.column_dimensions["G"].width=20
    sheet.column_dimensions["H"].width=25
    sheet.column_dimensions["I"].width=25
    sheet.column_dimensions["J"].width=27
    sheet.column_dimensions["K"].width=18
    sheet.column_dimensions["L"].width=18
    sheet.column_dimensions["M"].width=17
    sheet.column_dimensions["N"].width=17
    sheet.column_dimensions["O"].width=17
    sheet.column_dimensions["P"].width=17
    sheet.column_dimensions["Q"].width=17
    sheet.column_dimensions["R"].width=17
    sheet.column_dimensions["S"].width=17
    sheet.column_dimensions["T"].width=17
    
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    
    book.save(report_path)
    





def create_excel_by_local_business(report_path, report_type, locals, db_session):
    book=openpyxl.Workbook()
    sheet = book.active
    start_timer = time.time()
    get_by_copy = db_session.execute(text(f''' 
                                        SELECT 
                                            CR.NAME AS REGION,
                                            LC.CODE AS LOCAL_CODE,
                                            FT.ID AS FROM_TYPE_ID,
                                            SUM(BT.DEBT_ACCOUNT_START_STATE) as DEBT_ACCOUNT_START_STATE,
                                            SUM(CASE WHEN BT.DEBT_ACCOUNT_CREDIT_SUM IS NOT NULL THEN (BT.DEBT_ACCOUNT_CREDIT_SUM::real)/100 ELSE 0 END) as DEBT_ACCOUNT_CREDIT_SUM,
                                            SUM(
                                            CASE WHEN BT.DEBT_ACCOUNT_START_STATE IS NOT NULL THEN BT.DEBT_ACCOUNT_START_STATE::real ELSE 0 END
                                            -
                                            CASE WHEN BT.DEBT_ACCOUNT_CREDIT_SUM IS NOT NULL THEN (BT.DEBT_ACCOUNT_CREDIT_SUM::real)/100 ELSE 0 END) as DEBT_ACCOUNT_BALANCE,
                                            SUM(BT.ACCOUNT_16377_START_STATE) as ACCOUNT_16377_START_STATE,
                                            SUM(CASE WHEN BT.ACCOUNT_16377_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_DEBIT_SUM,
                                            SUM(CASE WHEN BT.ACCOUNT_16377_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_CREDIT_SUM,
                                            SUM((
                                            CASE WHEN BT.ACCOUNT_16377_START_STATE IS NOT NULL THEN BT.ACCOUNT_16377_START_STATE::real ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_16377_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_16377_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_BALANCE,
                                            
                                            SUM(CASE WHEN BT.ACCOUNT_163XX_START_STATE IS NOT NULL THEN (BT.ACCOUNT_163XX_START_STATE::real) ELSE 0 END) as ACCOUNT_163XX_START_STATE,
                                            SUM(CASE WHEN BT.ACCOUNT_163XX_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_DEBIT_SUM,
                                            SUM(CASE WHEN BT.ACCOUNT_163XX_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_CREDIT_SUM,
                                            SUM((CASE WHEN BT.ACCOUNT_163XX_START_STATE IS NOT NULL THEN BT.ACCOUNT_163XX_START_STATE::real ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_163XX_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_163XX_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_BALANCE
                                        FROM BALANCE_TURNOVER BT
                                        JOIN LOAN_PORTFOLIO LP ON BT.LOAN_ID = LP.LOAN_ID
                                        JOIN CLIENT_REGION CR ON LP.CLIENT_REGION = CR.ID
                                        JOIN LOCAL_CODE LC ON LP.LOCAL_CODE_ID = LC.ID
                                        {report_type}
                                        AND LP.LOCAL_CODE_ID {locals}
                                        GROUP BY LC.CODE, CR.NAME, FT.ID
                                          ''')).fetchall()

 
    print(time.time() - start_timer)
    row = 3
    i=1
    for loan in get_by_copy:
        i=i+1
        
        sheet.cell(column=1, row=row).value = i
        sheet.cell(column=2, row=row).value = loan['region']
        sheet.cell(column=3, row=row).value = loan['local_code']
        sheet.cell(column=4, row=row).value = loan['debt_account_start_state']
        if loan['from_type_id'] == 1:
            sheet.cell(column=5, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 2:
            sheet.cell(column=6, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 3:
            sheet.cell(column=7, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 4:
            sheet.cell(column=8, row=row).value = loan['debt_account_start_state']
        sheet.cell(column=9, row=row).value = loan['debt_account_credit_sum']
        sheet.cell(column=10, row=row).value = loan['debt_account_balance']
        sheet.cell(column=11, row=row).value = loan['account_16377_start_state']
        sheet.cell(column=12, row=row).value = loan['account_16377_debit_sum']
        sheet.cell(column=13, row=row).value = loan['account_16377_credit_sum']
        sheet.cell(column=14, row=row).value = loan['account_16377_balance']
        sheet.cell(column=15, row=row).value = loan['account_163xx_start_state']
        sheet.cell(column=16, row=row).value = loan['account_163xx_debit_sum']
        sheet.cell(column=17, row=row).value = loan['account_163xx_credit_sum']
        sheet.cell(column=18, row=row).value = loan['account_163xx_balance']
        
        row = row + 1
    
    #sheet.insert_rows(1)
    sheet.cell(column=5, row=1).value = "Остаток кредитов с просрочкой до 30 дней"
    sheet.cell(column=6, row=1).value = "Остаток кредитов с просрочкой до 30 дней из отчета за прошлый месяц"
    sheet.cell(column=7, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 31-60 дней"
    sheet.cell(column=8, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 60+  дней"
    
    sheet.insert_rows(1)
    sheet.cell(column=4, row=1).value = "Остаток кредитов к началу отчетного периода"
    sheet.cell(column=5, row=1).value = "В том числе"
    sheet.cell(column=9, row=1).value = "Погашено кредитов за отчетный период"
    sheet.cell(column=10, row=1).value = "Остаток кредитов к концу отчетного периода"
    
    sheet.insert_rows(1)
    sheet.cell(column=1, row=1).value = "ID"
    sheet.cell(column=2, row=1).value =  "Регион"
    sheet.cell(column=3, row=1).value = "Локал код"
    sheet.cell(column=4, row=1).value = "Задолженность с просрочкой до 30 дней (0-30)"
    sheet.cell(column=11, row=1).value = "Остаток 16377 на начало отчетного периода"
    sheet.cell(column=12, row=1).value = "Зачислено в течении месяца по счёту 16377"
    sheet.cell(column=13, row=1).value = "Погашено за отчетный период по счёту 16377"
    sheet.cell(column=14, row=1).value = "Остаток 16377 к концу отчетного периода"
    sheet.cell(column=15, row=1).value = "Остаток 163% на начало отчетного периода (кроме 16377)"
    sheet.cell(column=16, row=1).value = "Зачислено в течении месяца по счетам 163% (кроме 16377)"
    sheet.cell(column=17, row=1).value = "Погашено за отчетный период по счетам 163% (кроме 16377)"
    sheet.cell(column=18, row=1).value = "Остаток 163% к концу отчетного периода (кроме 16377)"
    
    bold = Font(bold=True)

    for row in sheet["1:3"]:
        for cell in row:
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    sheet.merge_cells('E2:H2')
    sheet.merge_cells('D1:J1')
    sheet.merge_cells('A1:A3')
    sheet.merge_cells('B1:B3')
    sheet.merge_cells('C1:C3')
    sheet.merge_cells('D2:D3')
    sheet.merge_cells('I2:I3')
    sheet.merge_cells('J2:J3')
    sheet.merge_cells('K1:K3')
    sheet.merge_cells('L1:L3')
    sheet.merge_cells('M1:M3')
    sheet.merge_cells('N1:N3')
    sheet.merge_cells('O1:O3')
    sheet.merge_cells('P1:P3')
    sheet.merge_cells('Q1:Q3')
    sheet.merge_cells('R1:R3')
    
    sheet.column_dimensions["A"].width=8
    sheet.column_dimensions["B"].width=8
    sheet.column_dimensions["C"].width=8
    sheet.column_dimensions["D"].width=20
    sheet.column_dimensions["E"].width=20
    sheet.column_dimensions["F"].width=25
    sheet.column_dimensions["G"].width=25
    sheet.column_dimensions["H"].width=27
    sheet.column_dimensions["I"].width=18
    sheet.column_dimensions["J"].width=18
    sheet.column_dimensions["K"].width=17
    sheet.column_dimensions["L"].width=17
    sheet.column_dimensions["M"].width=17
    sheet.column_dimensions["N"].width=17
    sheet.column_dimensions["O"].width=17
    sheet.column_dimensions["P"].width=17
    sheet.column_dimensions["Q"].width=17
    sheet.column_dimensions["R"].width=17
    
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    
    book.save(report_path)
    print('end: ', time.time() - start_timer, " sec")



def create_excel_by_client_kad(report_path, report_type, locals, db_session):
    book=openpyxl.Workbook()
    sheet = book.active
    get_by_copy = db_session.execute(text(f''' 
                                          SELECT BT.ID,
                                            CR.NAME AS REGION,
                                            LC.CODE AS LOCAL_CODE,
                                            LP.LOAN_ID,
                                            LP.CLIENT_NAME,
                                            FT.ID AS FROM_TYPE_ID,
                                            BT.DEBT_ACCOUNT_START_STATE,
                                            (CASE WHEN BT.DEBT_ACCOUNT_CREDIT_SUM IS NOT NULL THEN (BT.DEBT_ACCOUNT_CREDIT_SUM::real)/100 ELSE 0 END) as DEBT_ACCOUNT_CREDIT_SUM,
                                            (
                                            CASE WHEN BT.DEBT_ACCOUNT_START_STATE IS NOT NULL THEN BT.DEBT_ACCOUNT_START_STATE::real ELSE 0 END
                                            -
                                            CASE WHEN BT.DEBT_ACCOUNT_CREDIT_SUM IS NOT NULL THEN (BT.DEBT_ACCOUNT_CREDIT_SUM::real)/100 ELSE 0 END) as DEBT_ACCOUNT_BALANCE,
                                            BT.ACCOUNT_16377_START_STATE,
                                            (CASE WHEN BT.ACCOUNT_16377_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_DEBIT_SUM,
                                            (CASE WHEN BT.ACCOUNT_16377_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_CREDIT_SUM,
                                            ((
                                            CASE WHEN BT.ACCOUNT_16377_START_STATE IS NOT NULL THEN BT.ACCOUNT_16377_START_STATE::real ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_16377_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_16377_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_BALANCE,
                                            
                                            (CASE WHEN BT.ACCOUNT_163XX_START_STATE IS NOT NULL THEN (BT.ACCOUNT_163XX_START_STATE::real) ELSE 0 END) as ACCOUNT_163XX_START_STATE,
                                            (CASE WHEN BT.ACCOUNT_163XX_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_DEBIT_SUM,
                                            (CASE WHEN BT.ACCOUNT_163XX_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_CREDIT_SUM,
                                            
                                            ((CASE WHEN BT.ACCOUNT_163XX_START_STATE IS NOT NULL THEN (BT.ACCOUNT_163XX_START_STATE::real) ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_163XX_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_163XX_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_BALANCE,
                                            CASE WHEN HL1.GENERAL_TASK_ID = 7 THEN LS1.NAME ELSE '' END AS "35_letter",
                                            CASE WHEN HL1.GENERAL_TASK_ID = 7 THEN HL1.SEND_DATE ELSE NULL END AS SEND_DATE1,
                                            CASE WHEN HL1.GENERAL_TASK_ID = 7 THEN HL1.ERROR_COMMENT ELSE '' END AS "35_letter_error_comment",
                                            CASE WHEN HL2.GENERAL_TASK_ID = 8 THEN LS2.NAME ELSE '' END AS "45_letter",
                                            CASE WHEN HL2.GENERAL_TASK_ID = 8 THEN HL2.SEND_DATE ELSE NULL END AS SEND_DATE2,
                                            CASE WHEN HL2.GENERAL_TASK_ID = 8 THEN HL2.ERROR_COMMENT ELSE '' END AS "45_letter_error_comment"
                                        FROM BALANCE_TURNOVER BT
                                        LEFT JOIN LOAN_PORTFOLIO LP ON BT.LOAN_ID = LP.LOAN_ID
                                        JOIN CLIENT_REGION CR ON LP.CLIENT_REGION = CR.ID
                                        LEFT JOIN LOCAL_CODE LC ON LP.LOCAL_CODE_ID = LC.ID
                                        {report_type}
                                        AND LP.STATUS=1
                                        AND LP.LOCAL_CODE_ID {locals}
                                        
                                          '''))
 
    
    row = 3
    i=1
    
    for loan in get_by_copy:
        i=i+1
        
        sheet.cell(column=1, row=row).value = i
        sheet.cell(column=2, row=row).value = loan['region']
        sheet.cell(column=3, row=row).value = loan['local_code']
        sheet.cell(column=4, row=row).value = loan['loan_id']
        sheet.cell(column=5, row=row).value = loan['client_name']
        sheet.cell(column=6, row=row).value = loan['debt_account_start_state']
        if loan['from_type_id'] == 2:
            sheet.cell(column=7, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 3:
            sheet.cell(column=8, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 4:
            sheet.cell(column=9, row=row).value = loan['debt_account_start_state']
        sheet.cell(column=10, row=row).value = loan['debt_account_credit_sum']
        sheet.cell(column=11, row=row).value = loan['debt_account_balance']
        sheet.cell(column=12, row=row).value = loan['account_16377_start_state']
        sheet.cell(column=13, row=row).value = loan['account_16377_debit_sum']
        sheet.cell(column=14, row=row).value = loan['account_16377_credit_sum']
        sheet.cell(column=15, row=row).value = loan['account_16377_balance']
        sheet.cell(column=16, row=row).value = loan['account_163xx_start_state']
        sheet.cell(column=17, row=row).value = loan['account_163xx_debit_sum']
        sheet.cell(column=18, row=row).value = loan['account_163xx_credit_sum']
        sheet.cell(column=19, row=row).value = loan['account_163xx_balance']
        sheet.cell(column=20, row=row).value = loan['35_letter']
        sheet.cell(column=21, row=row).value = loan['send_date1']
        sheet.cell(column=22, row=row).value = loan['35_letter_error_comment']
        sheet.cell(column=23, row=row).value = loan['45_letter']
        sheet.cell(column=24, row=row).value = loan['send_date2']
        sheet.cell(column=25, row=row).value = loan['45_letter_error_comment']
        row = row + 1
    
    #sheet.insert_rows(1)
    sheet.cell(column=7, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 30 дней"
    sheet.cell(column=8, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 31-60 дней"
    sheet.cell(column=9, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 60+  дней"
    
    sheet.insert_rows(1)
    sheet.cell(column=6, row=1).value = "Остаток кредитов к началу отчетного периода"
    sheet.cell(column=7, row=1).value = "В том числе"
    sheet.cell(column=10, row=1).value = "Погашено кредитов за отчетный период"
    sheet.cell(column=11, row=1).value = "Остаток кредитов к концу отчетного периода"
    
    sheet.insert_rows(1)
    sheet.cell(column=1, row=1).value = "ID"
    sheet.cell(column=2, row=1).value =  "Регион"
    sheet.cell(column=3, row=1).value = "Локал код"
    sheet.cell(column=4, row=1).value = "ID кредита"
    sheet.cell(column=5, row=1).value = "Наименование клиента"
    sheet.cell(column=6, row=1).value = "Задолженность с просрочкой до 31-60 дней"
    sheet.cell(column=12, row=1).value = "Остаток 16377 на начало отчетного периода"
    sheet.cell(column=13, row=1).value = "Зачислено в течении месяца по счёту 16377"
    sheet.cell(column=14, row=1).value = "Погашено за отчетный период по счёту 16377"
    sheet.cell(column=15, row=1).value = "Остаток 16377 к концу отчетного периода"
    sheet.cell(column=16, row=1).value = "Остаток 163% на начало отчетного периода (кроме 16377)"
    sheet.cell(column=17, row=1).value = "Зачислено в течении месяца по счетам 163% (кроме 16377)"
    sheet.cell(column=18, row=1).value = "Погашено за отчетный период по счетам 163% (кроме 16377)"
    sheet.cell(column=19, row=1).value = "Остаток 163% к концу отчетного периода (кроме 16377)"
    sheet.cell(column=20, row=1).value = "Почта 35"
    sheet.cell(column=21, row=1).value = "Дата отправки"
    sheet.cell(column=22, row=1).value = "Ошибка почты 35"
    sheet.cell(column=23, row=1).value = "Почта 45"
    sheet.cell(column=24, row=1).value = "Дата отправки"
    sheet.cell(column=25, row=1).value = "Ошибка почты 45"
    
    bold = Font(bold=True)

    for row in sheet["1:3"]:
        for cell in row:
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    sheet.merge_cells('G2:I2')
    sheet.merge_cells('F1:K1')
    sheet.merge_cells('A1:A3')
    sheet.merge_cells('B1:B3')
    sheet.merge_cells('C1:C3')
    sheet.merge_cells('D1:D3')
    sheet.merge_cells('E1:E3')
    sheet.merge_cells('F2:F3')
    sheet.merge_cells('J2:J3')
    sheet.merge_cells('K2:K3')
    sheet.merge_cells('L1:L3')
    sheet.merge_cells('M1:M3')
    sheet.merge_cells('N1:N3')
    sheet.merge_cells('O1:O3')
    sheet.merge_cells('P1:P3')
    sheet.merge_cells('Q1:Q3')
    sheet.merge_cells('R1:R3')
    sheet.merge_cells('S1:S3')
    sheet.merge_cells('T1:T3')
    sheet.merge_cells('U1:U3')
    sheet.merge_cells('V1:V3')
    sheet.merge_cells('W1:W3')
    sheet.merge_cells('X1:X3')
    sheet.merge_cells('Y1:Y3')
    
    sheet.column_dimensions["A"].width=8
    sheet.column_dimensions["B"].width=8
    sheet.column_dimensions["C"].width=8
    sheet.column_dimensions["D"].width=11
    sheet.column_dimensions["E"].width=33
    sheet.column_dimensions["F"].width=20
    sheet.column_dimensions["G"].width=25
    sheet.column_dimensions["H"].width=25
    sheet.column_dimensions["I"].width=25
    sheet.column_dimensions["J"].width=18
    sheet.column_dimensions["K"].width=18
    sheet.column_dimensions["L"].width=17
    sheet.column_dimensions["M"].width=17
    sheet.column_dimensions["N"].width=17
    sheet.column_dimensions["O"].width=17
    sheet.column_dimensions["P"].width=17
    sheet.column_dimensions["Q"].width=17
    sheet.column_dimensions["R"].width=17
    sheet.column_dimensions["S"].width=17
    sheet.column_dimensions["T"].width=17
    sheet.column_dimensions["U"].width=17
    sheet.column_dimensions["V"].width=17
    sheet.column_dimensions["W"].width=17
    sheet.column_dimensions["X"].width=17
    sheet.column_dimensions["Y"].width=17
    
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    
    book.save(report_path)
    





def create_excel_by_local_kad(report_path, report_type, locals, db_session):
    book=openpyxl.Workbook()
    sheet = book.active
    start_timer = time.time()
    get_by_copy = db_session.execute(text(f''' 
                                        SELECT 
                                            CR.NAME AS REGION,
                                            LC.CODE AS LOCAL_CODE,
                                            FT.ID AS FROM_TYPE_ID,
                                            SUM(BT.DEBT_ACCOUNT_START_STATE) as DEBT_ACCOUNT_START_STATE,
                                            SUM(CASE WHEN BT.DEBT_ACCOUNT_CREDIT_SUM IS NOT NULL THEN (BT.DEBT_ACCOUNT_CREDIT_SUM::real)/100 ELSE 0 END) as DEBT_ACCOUNT_CREDIT_SUM,
                                            SUM(
                                            CASE WHEN BT.DEBT_ACCOUNT_START_STATE IS NOT NULL THEN BT.DEBT_ACCOUNT_START_STATE::real ELSE 0 END
                                            -
                                            CASE WHEN BT.DEBT_ACCOUNT_CREDIT_SUM IS NOT NULL THEN (BT.DEBT_ACCOUNT_CREDIT_SUM::real)/100 ELSE 0 END) as DEBT_ACCOUNT_BALANCE,
                                            SUM(BT.ACCOUNT_16377_START_STATE) as ACCOUNT_16377_START_STATE,
                                            SUM(CASE WHEN BT.ACCOUNT_16377_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_DEBIT_SUM,
                                            SUM(CASE WHEN BT.ACCOUNT_16377_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_CREDIT_SUM,
                                            SUM((
                                            CASE WHEN BT.ACCOUNT_16377_START_STATE IS NOT NULL THEN BT.ACCOUNT_16377_START_STATE::real ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_16377_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_16377_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_BALANCE,
                                            
                                            SUM(CASE WHEN BT.ACCOUNT_163XX_START_STATE IS NOT NULL THEN (BT.ACCOUNT_163XX_START_STATE::real) ELSE 0 END) as ACCOUNT_163XX_START_STATE,
                                            SUM(CASE WHEN BT.ACCOUNT_163XX_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_DEBIT_SUM,
                                            SUM(CASE WHEN BT.ACCOUNT_163XX_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_CREDIT_SUM,
                                            SUM((CASE WHEN BT.ACCOUNT_163XX_START_STATE IS NOT NULL THEN BT.ACCOUNT_163XX_START_STATE::real ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_163XX_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_163XX_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_BALANCE
                                        FROM BALANCE_TURNOVER BT
                                        JOIN LOAN_PORTFOLIO LP ON BT.LOAN_ID = LP.LOAN_ID
                                        JOIN CLIENT_REGION CR ON LP.CLIENT_REGION = CR.ID
                                        JOIN LOCAL_CODE LC ON LP.LOCAL_CODE_ID = LC.ID
                                        {report_type}
                                        AND LP.LOCAL_CODE_ID {locals}
                                        GROUP BY LC.CODE, CR.NAME, FT.ID
                                          ''')).fetchall()

 
    print(time.time() - start_timer)
    row = 3
    i=1
    for loan in get_by_copy:
        i=i+1
        
        sheet.cell(column=1, row=row).value = i
        sheet.cell(column=2, row=row).value = loan['region']
        sheet.cell(column=3, row=row).value = loan['local_code']
        sheet.cell(column=4, row=row).value = loan['debt_account_start_state']
        if loan['from_type_id'] == 2:
            sheet.cell(column=5, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 3:
            sheet.cell(column=6, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 4:
            sheet.cell(column=7, row=row).value = loan['debt_account_start_state']
        sheet.cell(column=8, row=row).value = loan['debt_account_credit_sum']
        sheet.cell(column=9, row=row).value = loan['debt_account_balance']
        sheet.cell(column=10, row=row).value = loan['account_16377_start_state']
        sheet.cell(column=11, row=row).value = loan['account_16377_debit_sum']
        sheet.cell(column=12, row=row).value = loan['account_16377_credit_sum']
        sheet.cell(column=13, row=row).value = loan['account_16377_balance']
        sheet.cell(column=14, row=row).value = loan['account_163xx_start_state']
        sheet.cell(column=15, row=row).value = loan['account_163xx_debit_sum']
        sheet.cell(column=16, row=row).value = loan['account_163xx_credit_sum']
        sheet.cell(column=17, row=row).value = loan['account_163xx_balance']
        
        row = row + 1
    
    #sheet.insert_rows(1)
    sheet.cell(column=5, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 30 дней"
    sheet.cell(column=6, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 31-60 дней"
    sheet.cell(column=7, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 60+  дней"
    
    sheet.insert_rows(1)
    sheet.cell(column=4, row=1).value = "Остаток кредитов к началу отчетного периода"
    sheet.cell(column=5, row=1).value = "В том числе"
    sheet.cell(column=8, row=1).value = "Погашено кредитов за отчетный период"
    sheet.cell(column=9, row=1).value = "Остаток кредитов к концу отчетного периода"
    
    sheet.insert_rows(1)
    sheet.cell(column=1, row=1).value = "ID"
    sheet.cell(column=2, row=1).value =  "Регион"
    sheet.cell(column=3, row=1).value = "Локал код"
    sheet.cell(column=4, row=1).value = "Задолженность с просрочкой до 31-60 дней"
    sheet.cell(column=10, row=1).value = "Остаток 16377 на начало отчетного периода"
    sheet.cell(column=11, row=1).value = "Зачислено в течении месяца по счёту 16377"
    sheet.cell(column=12, row=1).value = "Погашено за отчетный период по счёту 16377"
    sheet.cell(column=13, row=1).value = "Остаток 16377 к концу отчетного периода"
    sheet.cell(column=14, row=1).value = "Остаток 163% на начало отчетного периода (кроме 16377)"
    sheet.cell(column=15, row=1).value = "Зачислено в течении месяца по счетам 163% (кроме 16377)"
    sheet.cell(column=16, row=1).value = "Погашено за отчетный период по счетам 163% (кроме 16377)"
    sheet.cell(column=17, row=1).value = "Остаток 163% к концу отчетного периода (кроме 16377)"
    
    bold = Font(bold=True)

    for row in sheet["1:3"]:
        for cell in row:
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    sheet.merge_cells('E2:G2')
    sheet.merge_cells('D1:I1')
    sheet.merge_cells('A1:A3')
    sheet.merge_cells('B1:B3')
    sheet.merge_cells('C1:C3')
    sheet.merge_cells('D2:D3')
    sheet.merge_cells('H2:H3')
    sheet.merge_cells('I2:I3')
    sheet.merge_cells('J1:J3')
    sheet.merge_cells('K1:K3')
    sheet.merge_cells('L1:L3')
    sheet.merge_cells('M1:M3')
    sheet.merge_cells('N1:N3')
    sheet.merge_cells('O1:O3')
    sheet.merge_cells('P1:P3')
    sheet.merge_cells('Q1:Q3')
    
    sheet.column_dimensions["A"].width=8
    sheet.column_dimensions["B"].width=8
    sheet.column_dimensions["C"].width=8
    sheet.column_dimensions["D"].width=20
    sheet.column_dimensions["E"].width=25
    sheet.column_dimensions["F"].width=25
    sheet.column_dimensions["G"].width=25
    sheet.column_dimensions["H"].width=18
    sheet.column_dimensions["I"].width=18
    sheet.column_dimensions["J"].width=17
    sheet.column_dimensions["K"].width=17
    sheet.column_dimensions["L"].width=17
    sheet.column_dimensions["M"].width=17
    sheet.column_dimensions["N"].width=17
    sheet.column_dimensions["O"].width=17
    sheet.column_dimensions["P"].width=17
    sheet.column_dimensions["Q"].width=17
    
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    
    book.save(report_path)
    print('end: ', time.time() - start_timer, " sec")



def create_excel_by_client_problems(report_path, report_type, locals, db_session):
    book=openpyxl.Workbook()
    sheet = book.active
    get_by_copy = db_session.execute(text(f''' 
                                          SELECT BT.ID,
                                            PC.ID AS PROBLEMS_CASE_ID,
                                            CR.NAME AS REGION,
                                            LC.CODE AS LOCAL_CODE,
                                            LP.LOAN_ID,
                                            LP.ID AS LOAN_PORTFOLIO_ID,
                                            LP.CLIENT_NAME,
                                            FT.ID AS FROM_TYPE_ID,
                                            BT.DEBT_ACCOUNT_START_STATE,
                                            (CASE WHEN BT.DEBT_ACCOUNT_CREDIT_SUM IS NOT NULL THEN (BT.DEBT_ACCOUNT_CREDIT_SUM::real)/100 ELSE 0 END) as DEBT_ACCOUNT_CREDIT_SUM,
                                            (
                                            CASE WHEN BT.DEBT_ACCOUNT_START_STATE IS NOT NULL THEN BT.DEBT_ACCOUNT_START_STATE::real ELSE 0 END
                                            -
                                            CASE WHEN BT.DEBT_ACCOUNT_CREDIT_SUM IS NOT NULL THEN (BT.DEBT_ACCOUNT_CREDIT_SUM::real)/100 ELSE 0 END) as DEBT_ACCOUNT_BALANCE,
                                            BT.ACCOUNT_16377_START_STATE,
                                            (CASE WHEN BT.ACCOUNT_16377_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_DEBIT_SUM,
                                            (CASE WHEN BT.ACCOUNT_16377_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_CREDIT_SUM,
                                            ((
                                            CASE WHEN BT.ACCOUNT_16377_START_STATE IS NOT NULL THEN BT.ACCOUNT_16377_START_STATE::real ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_16377_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_16377_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_BALANCE,
                                            
                                            (CASE WHEN BT.ACCOUNT_163XX_START_STATE IS NOT NULL THEN (BT.ACCOUNT_163XX_START_STATE::real) ELSE 0 END) as ACCOUNT_163XX_START_STATE,
                                            (CASE WHEN BT.ACCOUNT_163XX_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_DEBIT_SUM,
                                            (CASE WHEN BT.ACCOUNT_163XX_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_CREDIT_SUM,
                                            
                                            ((CASE WHEN BT.ACCOUNT_163XX_START_STATE IS NOT NULL THEN (BT.ACCOUNT_163XX_START_STATE::real) ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_163XX_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_163XX_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_BALANCE,
                                            PS.NAME,
                                            PS.ID AS PS_ID
                                        FROM BALANCE_TURNOVER BT
                                        LEFT JOIN LOAN_PORTFOLIO LP ON BT.LOAN_ID = LP.LOAN_ID
                                        JOIN CLIENT_REGION CR ON LP.CLIENT_REGION = CR.ID
                                        LEFT JOIN PROBLEM_STATE_CHAIN PSCH ON PSCH.LOAN_ID = LP.LOAN_ID
                                        LEFT JOIN PROBLEM_STATES PS ON PS.ID = PSCH.LAST_STATE_ID
                                        LEFT JOIN LOCAL_CODE LC ON LP.LOCAL_CODE_ID = LC.ID
                                        {report_type}
                                        AND LP.STATUS=1
                                        AND LP.LOCAL_CODE_ID {locals}
                                        
                                          ''')).fetchall()

    
    row = 3
    i=1
    
    for loan in get_by_copy:
        i=i+1
        set_status = None
        set_date = None
        if loan.ps_id is not None:
            
            if loan.ps_id == 4:
                get_status = db_session.execute(text(f'''SELECT
                                                        JT.NAME_FULL AS NAME,
                                                        (CASE 
                                                        WHEN JD.TYPE_ID=1 THEN JD.RECEIPT_DATE
                                                        WHEN JD.TYPE_ID=2 THEN JD.DECISION_DATE_ON_ADMISSION 
                                                        WHEN JD.TYPE_ID=3 THEN JD.DECISION_DATE_TO_SET
                                                        WHEN JD.TYPE_ID=4 THEN JD.DECISION_DATE_IN_FAVOR_OF_BANK
                                                        ELSE null END)  JUDICIAL_DATE
                                                            FROM
                                                        JUDICIAL_DATA JD
                                                        JOIN JUDICIAL_TYPE JT ON JT.ID = JD.TYPE_ID
                                                        JOIN JUDICIAL_AUTHORITY JA ON JA.ID = JD.AUTHORITY_ID
                                                        JOIN JUDICIAL_AUTHORITY_TYPE JAT ON JAT.ID = JA.TYPE_ID
                                                        WHERE JD.PROBLEMS_CASE_ID={loan.problems_case_id}''')).fetchone()
                set_status = get_status and get_status.name or None
                set_date = get_status and get_status.judicial_date.date() or None
            if loan.ps_id == 5: 
                get_status = db_session.execute(text(f'''SELECT
                                                        IL.RESPONSE
                                                        FROM
                                                        INTEGRATIONS_LOG IL
                                                        WHERE LOAN_PORTFOLIO_ID={loan.loan_portfolio_id}
                                                        AND (SERVICE_API_ID=18 OR SERVICE_API_ID=19)
                                                        ''')).fetchone()
                if get_status is not None and 'action' in get_status[0]:
                    
                    set_status = get_status[0]['action'][0]['decision']
                    set_date = get_status[0]['action'][0]['decision_date']
                
                
                
            elif loan.ps_id == 6:
                get_status = db_session.execute(text(f'''SELECT
                                                        PAT.NAME_FULL AS NAME
                                                        FROM
                                                        PROBLEMS_ASSETS PA
                                                        JOIN PROBLEMS_ASSETS_TYPE PAT ON PAT.id = PA.type_id
                                                        WHERE PA.PROBLEMS_CASE_ID={loan.problems_case_id}''')).fetchone()
                set_status = get_status and get_status.name or None
            
            elif loan.ps_id == 7:
                set_status = 'Списание кредитов, лизинга и факторинга'
            
            elif loan.ps_id == 8:
                get_status = db_session.execute(text(f'''SELECT
                                                        PAT.NAME
                                                        FROM
                                                        PROBLEMS_AUCTION PA
                                                        JOIN PROBLEMS_AUCTION_TYPE PAT ON PAT.id = PA.type_id
                                                        WHERE PA.PROBLEMS_CASE_ID={loan.problems_case_id}''')).fetchone()
                set_status = get_status and get_status.name or None
                
                
            elif loan.ps_id == 9:
                get_status = db_session.execute(text(f'''SELECT
                                                        PMT.NAME
                                                        FROM
                                                        PROBLEMS_MIB PM
                                                        JOIN PROBLEMS_MIB_TYPE PMT ON PMT.id = PM.type_id
                                                        WHERE PM.PROBLEMS_CASE_ID={loan.problems_case_id}''')).fetchone()
                set_status = get_status and get_status.name or None
                
                
        sheet.cell(column=1, row=row).value = i
        sheet.cell(column=2, row=row).value = loan['region']
        sheet.cell(column=3, row=row).value = loan['local_code']
        sheet.cell(column=4, row=row).value = loan['loan_id']
        sheet.cell(column=5, row=row).value = loan['client_name']
        sheet.cell(column=6, row=row).value = loan['name']
        sheet.cell(column=7, row=row).value = set_status
        sheet.cell(column=8, row=row).value = set_date
        sheet.cell(column=9, row=row).value = loan['debt_account_start_state']
        if loan['from_type_id'] == 2:
            sheet.cell(column=10, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 3:
            sheet.cell(column=11, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 4:
            sheet.cell(column=12, row=row).value = loan['debt_account_start_state']
        sheet.cell(column=13, row=row).value = loan['debt_account_credit_sum']
        sheet.cell(column=14, row=row).value = loan['debt_account_balance']
        sheet.cell(column=15, row=row).value = loan['account_16377_start_state']
        sheet.cell(column=16, row=row).value = loan['account_16377_debit_sum']
        sheet.cell(column=17, row=row).value = loan['account_16377_credit_sum']
        sheet.cell(column=18, row=row).value = loan['account_16377_balance']
        sheet.cell(column=19, row=row).value = loan['account_163xx_start_state']
        sheet.cell(column=20, row=row).value = loan['account_163xx_debit_sum']
        sheet.cell(column=21, row=row).value = loan['account_163xx_credit_sum']
        sheet.cell(column=22, row=row).value = loan['account_163xx_balance']
        
        row = row + 1
    
    #sheet.insert_rows(1)
    sheet.cell(column=10, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 30 дней"
    sheet.cell(column=11, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 31-60 дней"
    sheet.cell(column=12, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 60+  дней"
    
    sheet.insert_rows(1)
    sheet.cell(column=9, row=1).value = "Остаток кредитов к началу отчетного периода"
    sheet.cell(column=10, row=1).value = "В том числе"
    sheet.cell(column=13, row=1).value = "Погашено кредитов за отчетный период"
    sheet.cell(column=14, row=1).value = "Остаток кредитов к концу отчетного периода"
    
    sheet.insert_rows(1)
    sheet.cell(column=1, row=1).value = "ID"
    sheet.cell(column=2, row=1).value =  "Регион"
    sheet.cell(column=3, row=1).value = "Локал код"
    sheet.cell(column=4, row=1).value = "ID кредита"
    sheet.cell(column=5, row=1).value = "Наименование клиента"
    sheet.cell(column=6, row=1).value = "Статус кредита"
    sheet.cell(column=7, row=1).value = "Описание статуса кредита"
    sheet.cell(column=8, row=1).value = "Дата статуса кредита"
    sheet.cell(column=9, row=1).value = "Задолженность с просрочкой до 31-60 дней"
    sheet.cell(column=15, row=1).value = "Остаток 16377 на начало отчетного периода"
    sheet.cell(column=16, row=1).value = "Зачислено в течении месяца по счёту 16377"
    sheet.cell(column=17, row=1).value = "Погашено за отчетный период по счёту 16377"
    sheet.cell(column=18, row=1).value = "Остаток 16377 к концу отчетного периода"
    sheet.cell(column=19, row=1).value = "Остаток 163% на начало отчетного периода (кроме 16377)"
    sheet.cell(column=20, row=1).value = "Зачислено в течении месяца по счетам 163% (кроме 16377)"
    sheet.cell(column=21, row=1).value = "Погашено за отчетный период по счетам 163% (кроме 16377)"
    sheet.cell(column=22, row=1).value = "Остаток 163% к концу отчетного периода (кроме 16377)"
    
    bold = Font(bold=True)

    for row in sheet["1:3"]:
        for cell in row:
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    sheet.merge_cells('J2:L2')
    sheet.merge_cells('I1:N1')
    sheet.merge_cells('A1:A3')
    sheet.merge_cells('B1:B3')
    sheet.merge_cells('C1:C3')
    sheet.merge_cells('D1:D3')
    sheet.merge_cells('E1:E3')
    sheet.merge_cells('F1:F3')
    sheet.merge_cells('G1:G3')
    sheet.merge_cells('H1:H3')
    sheet.merge_cells('I2:I3')
    sheet.merge_cells('M2:M3')
    sheet.merge_cells('N2:N3')
    sheet.merge_cells('O1:O3')
    sheet.merge_cells('P1:P3')
    sheet.merge_cells('Q1:Q3')
    sheet.merge_cells('R1:R3')
    sheet.merge_cells('S1:S3')
    sheet.merge_cells('T1:T3')
    sheet.merge_cells('U1:U3')
    sheet.merge_cells('V1:V3')
    
    sheet.column_dimensions["A"].width=8
    sheet.column_dimensions["B"].width=8
    sheet.column_dimensions["C"].width=8
    sheet.column_dimensions["D"].width=11
    sheet.column_dimensions["E"].width=33
    sheet.column_dimensions["F"].width=20
    sheet.column_dimensions["G"].width=20
    sheet.column_dimensions["H"].width=20
    sheet.column_dimensions["I"].width=25
    sheet.column_dimensions["J"].width=25
    sheet.column_dimensions["K"].width=18
    sheet.column_dimensions["L"].width=18
    sheet.column_dimensions["M"].width=17
    sheet.column_dimensions["N"].width=17
    sheet.column_dimensions["O"].width=17
    sheet.column_dimensions["P"].width=17
    sheet.column_dimensions["Q"].width=17
    sheet.column_dimensions["R"].width=17
    sheet.column_dimensions["S"].width=17
    sheet.column_dimensions["T"].width=17
    sheet.column_dimensions["U"].width=17
    sheet.column_dimensions["V"].width=17
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    
    book.save(report_path)
    





def create_excel_by_local_problems(report_path, report_type, locals, db_session):
    book=openpyxl.Workbook()
    sheet = book.active
    start_timer = time.time()
    get_by_copy = db_session.execute(text(f''' 
                                        SELECT 
                                            LC.CODE AS LOCAL_CODE,
                                            CR.NAME AS REGION,
                                            FT.ID AS FROM_TYPE_ID,
                                            SUM(BT.DEBT_ACCOUNT_START_STATE) as DEBT_ACCOUNT_START_STATE,
                                            SUM(CASE WHEN BT.DEBT_ACCOUNT_CREDIT_SUM IS NOT NULL THEN (BT.DEBT_ACCOUNT_CREDIT_SUM::real)/100 ELSE 0 END) as DEBT_ACCOUNT_CREDIT_SUM,
                                            SUM(
                                            CASE WHEN BT.DEBT_ACCOUNT_START_STATE IS NOT NULL THEN BT.DEBT_ACCOUNT_START_STATE::real ELSE 0 END
                                            -
                                            CASE WHEN BT.DEBT_ACCOUNT_CREDIT_SUM IS NOT NULL THEN (BT.DEBT_ACCOUNT_CREDIT_SUM::real)/100 ELSE 0 END) as DEBT_ACCOUNT_BALANCE,
                                            SUM(BT.ACCOUNT_16377_START_STATE) as ACCOUNT_16377_START_STATE,
                                            SUM(CASE WHEN BT.ACCOUNT_16377_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_DEBIT_SUM,
                                            SUM(CASE WHEN BT.ACCOUNT_16377_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_CREDIT_SUM,
                                            SUM((
                                            CASE WHEN BT.ACCOUNT_16377_START_STATE IS NOT NULL THEN BT.ACCOUNT_16377_START_STATE::real ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_16377_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_16377_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_16377_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_16377_BALANCE,
                                            
                                            SUM(CASE WHEN BT.ACCOUNT_163XX_START_STATE IS NOT NULL THEN (BT.ACCOUNT_163XX_START_STATE::real) ELSE 0 END) as ACCOUNT_163XX_START_STATE,
                                            SUM(CASE WHEN BT.ACCOUNT_163XX_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_DEBIT_SUM,
                                            SUM(CASE WHEN BT.ACCOUNT_163XX_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_CREDIT_SUM,
                                            SUM((CASE WHEN BT.ACCOUNT_163XX_START_STATE IS NOT NULL THEN BT.ACCOUNT_163XX_START_STATE::real ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_163XX_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_163XX_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_163XX_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_163XX_BALANCE
                                        FROM BALANCE_TURNOVER BT
                                        JOIN LOAN_PORTFOLIO LP ON BT.LOAN_ID = LP.LOAN_ID
                                        JOIN CLIENT_REGION CR ON LP.CLIENT_REGION = CR.ID
                                        JOIN LOCAL_CODE LC ON LP.LOCAL_CODE_ID = LC.ID
                                        {report_type}
                                        AND LP.LOCAL_CODE_ID {locals}
                                        GROUP BY LC.CODE, CR.NAME, FT.ID
                                          ''')).fetchall()

 
    print(time.time() - start_timer)
    row = 3
    i=1
    for loan in get_by_copy:
        i=i+1
        
        sheet.cell(column=1, row=row).value = i
        sheet.cell(column=2, row=row).value = loan['region']
        sheet.cell(column=3, row=row).value = loan['local_code']
        sheet.cell(column=4, row=row).value = loan['debt_account_start_state']
        if loan['from_type_id'] == 2:
            sheet.cell(column=5, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 3:
            sheet.cell(column=6, row=row).value = loan['debt_account_start_state']
        elif loan['from_type_id'] == 4:
            sheet.cell(column=7, row=row).value = loan['debt_account_start_state']
        sheet.cell(column=8, row=row).value = loan['debt_account_credit_sum']
        sheet.cell(column=9, row=row).value = loan['debt_account_balance']
        sheet.cell(column=10, row=row).value = loan['account_16377_start_state']
        sheet.cell(column=11, row=row).value = loan['account_16377_debit_sum']
        sheet.cell(column=12, row=row).value = loan['account_16377_credit_sum']
        sheet.cell(column=13, row=row).value = loan['account_16377_balance']
        sheet.cell(column=14, row=row).value = loan['account_163xx_start_state']
        sheet.cell(column=15, row=row).value = loan['account_163xx_debit_sum']
        sheet.cell(column=16, row=row).value = loan['account_163xx_credit_sum']
        sheet.cell(column=17, row=row).value = loan['account_163xx_balance']
        
        row = row + 1
    
    #sheet.insert_rows(1)
    sheet.cell(column=5, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 30 дней"
    sheet.cell(column=6, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 31-60 дней"
    sheet.cell(column=7, row=1).value = "Перенесено за счет взыскания задолженности прошлого месяца со сроком просрочки 60+  дней"
    
    sheet.insert_rows(1)
    sheet.cell(column=4, row=1).value = "Остаток кредитов к началу отчетного периода"
    sheet.cell(column=5, row=1).value = "В том числе"
    sheet.cell(column=8, row=1).value = "Погашено кредитов за отчетный период"
    sheet.cell(column=9, row=1).value = "Остаток кредитов к концу отчетного периода"
    
    sheet.insert_rows(1)
    sheet.cell(column=1, row=1).value = "ID"
    sheet.cell(column=2, row=1).value =  "Регион"
    sheet.cell(column=3, row=1).value = "Локал код"
    sheet.cell(column=4, row=1).value = "Задолженность с просрочкой до 60+ дней"
    sheet.cell(column=10, row=1).value = "Остаток 16377 на начало отчетного периода"
    sheet.cell(column=11, row=1).value = "Зачислено в течении месяца по счёту 16377"
    sheet.cell(column=12, row=1).value = "Погашено за отчетный период по счёту 16377"
    sheet.cell(column=13, row=1).value = "Остаток 16377 к концу отчетного периода"
    sheet.cell(column=14, row=1).value = "Остаток 163% на начало отчетного периода (кроме 16377)"
    sheet.cell(column=15, row=1).value = "Зачислено в течении месяца по счетам 163% (кроме 16377)"
    sheet.cell(column=16, row=1).value = "Погашено за отчетный период по счетам 163% (кроме 16377)"
    sheet.cell(column=17, row=1).value = "Остаток 163% к концу отчетного периода (кроме 16377)"
    
    bold = Font(bold=True)

    for row in sheet["1:3"]:
        for cell in row:
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    sheet.merge_cells('E2:G2')
    sheet.merge_cells('D1:I1')
    sheet.merge_cells('A1:A3')
    sheet.merge_cells('B1:B3')
    sheet.merge_cells('C1:C3')
    sheet.merge_cells('D2:D3')
    sheet.merge_cells('H2:H3')
    sheet.merge_cells('I2:I3')
    sheet.merge_cells('J1:J3')
    sheet.merge_cells('K1:K3')
    sheet.merge_cells('L1:L3')
    sheet.merge_cells('M1:M3')
    sheet.merge_cells('N1:N3')
    sheet.merge_cells('O1:O3')
    sheet.merge_cells('P1:P3')
    sheet.merge_cells('Q1:Q3')
    
    sheet.column_dimensions["A"].width=8
    sheet.column_dimensions["B"].width=8
    sheet.column_dimensions["C"].width=8
    sheet.column_dimensions["D"].width=20
    sheet.column_dimensions["E"].width=25
    sheet.column_dimensions["F"].width=25
    sheet.column_dimensions["G"].width=25
    sheet.column_dimensions["H"].width=18
    sheet.column_dimensions["I"].width=18
    sheet.column_dimensions["J"].width=17
    sheet.column_dimensions["K"].width=17
    sheet.column_dimensions["L"].width=17
    sheet.column_dimensions["M"].width=17
    sheet.column_dimensions["N"].width=17
    sheet.column_dimensions["O"].width=17
    sheet.column_dimensions["P"].width=17
    sheet.column_dimensions["Q"].width=17
    
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    
    book.save(report_path)
    print('end: ', time.time() - start_timer, " sec")



def create_excel_by_client_out_of_balance(report_path, report_type, locals, db_session):
    book=openpyxl.Workbook()
    sheet = book.active
    get_by_copy = db_session.execute(text(f''' 
                                          SELECT BT.ID,
                                            CR.NAME AS REGION,
                                            LC.CODE AS LOCAL_CODE,
                                            LP.LOAN_ID,
                                            LP.CLIENT_NAME,
                                            FT.ID AS FROM_TYPE_ID,
                                            (CASE WHEN BT.ACCOUNT_95413_START_STATE IS NOT NULL THEN BT.ACCOUNT_95413_START_STATE::real ELSE 0 END) as ACCOUNT_95413_START_STATE,
                                            (CASE WHEN BT.ACCOUNT_95413_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_95413_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_95413_DEBIT_SUM,
                                            (CASE WHEN BT.ACCOUNT_95413_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_95413_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_95413_CREDIT_SUM,
                                            (
                                            CASE WHEN BT.ACCOUNT_95413_START_STATE IS NOT NULL THEN BT.ACCOUNT_95413_START_STATE::real ELSE 0 END
                                            -
                                            CASE WHEN BT.ACCOUNT_95413_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_95413_CREDIT_SUM::real)/100 ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_95413_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_95413_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_95413_BALANCE,

                                            (CASE WHEN BT.ACCOUNT_9150X_START_STATE IS NOT NULL THEN (BT.ACCOUNT_9150X_START_STATE::real) ELSE 0 END) as ACCOUNT_9150X_START_STATE,
                                            (CASE WHEN BT.ACCOUNT_9150X_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_9150X_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_9150X_DEBIT_SUM,
                                            (CASE WHEN BT.ACCOUNT_9150X_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_9150X_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_9150X_CREDIT_SUM,

                                            ((CASE WHEN BT.ACCOUNT_9150X_START_STATE IS NOT NULL THEN (BT.ACCOUNT_9150X_START_STATE::real) ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_9150X_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_9150X_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_9150X_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_9150X_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_9150X_BALANCE
                                        FROM BALANCE_TURNOVER BT
                                        LEFT JOIN LOAN_PORTFOLIO LP ON BT.LOAN_ID = LP.LOAN_ID
                                        JOIN CLIENT_REGION CR ON LP.CLIENT_REGION = CR.ID
                                        LEFT JOIN LOCAL_CODE LC ON LP.LOCAL_CODE_ID = LC.ID
                                        {report_type}
                                        AND LP.STATUS=1
                                        AND LP.LOCAL_CODE_ID {locals}
                                        
                                          ''')).fetchall()

    
    row = 2
    i=1
    
    for loan in get_by_copy:
        i=i+1
        
        sheet.cell(column=1, row=row).value = i
        sheet.cell(column=2, row=row).value = loan['region']
        sheet.cell(column=3, row=row).value = loan['local_code']
        sheet.cell(column=4, row=row).value = loan['loan_id']
        sheet.cell(column=5, row=row).value = loan['client_name']
        sheet.cell(column=6, row=row).value = loan['account_95413_start_state']
        sheet.cell(column=7, row=row).value = loan['account_95413_debit_sum']
        sheet.cell(column=8, row=row).value = loan['account_95413_credit_sum']
        sheet.cell(column=9, row=row).value = loan['account_95413_balance']
        sheet.cell(column=10, row=row).value = loan['account_9150x_start_state']
        sheet.cell(column=11, row=row).value = loan['account_9150x_debit_sum']
        sheet.cell(column=12, row=row).value = loan['account_9150x_credit_sum']
        sheet.cell(column=13, row=row).value = loan['account_9150x_balance']
        
        row = row + 1
    
    #sheet.insert_rows(1)
    sheet.cell(column=1, row=1).value = "ID"
    sheet.cell(column=2, row=1).value =  "Регион"
    sheet.cell(column=3, row=1).value = "Локал код"
    sheet.cell(column=4, row=1).value = "ID кредита"
    sheet.cell(column=5, row=1).value = "Наименование клиента"
    sheet.cell(column=6, row=1).value = "Остаток 95413 на начало отчетного периода"
    sheet.cell(column=7, row=1).value = "Зачислено в течении месяца по счёту 95413"
    sheet.cell(column=8, row=1).value = "Погашено за отчетный период по счёту 95413"
    sheet.cell(column=9, row=1).value = "Остаток 95413 к концу отчетного периода"
    sheet.cell(column=10, row=1).value = "Остаток 91501 и 91503 на начало отчетного периода"
    sheet.cell(column=11, row=1).value = "Зачислено в течении месяца по счетам 91501 и 91503"
    sheet.cell(column=12, row=1).value = "Погашено за отчетный период по счетам 91501 и 91503"
    sheet.cell(column=13, row=1).value = "Остаток 91501 и 91503 к концу отчетного периода"
    
    bold = Font(bold=True)

    
    for cell in sheet["1:1"]:
        cell.font = bold
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    sheet.column_dimensions["A"].width=8
    sheet.column_dimensions["B"].width=8
    sheet.column_dimensions["C"].width=8
    sheet.column_dimensions["D"].width=11
    sheet.column_dimensions["E"].width=33
    sheet.column_dimensions["F"].width=20
    sheet.column_dimensions["G"].width=20
    sheet.column_dimensions["H"].width=20
    sheet.column_dimensions["I"].width=20
    sheet.column_dimensions["J"].width=20
    sheet.column_dimensions["K"].width=20
    sheet.column_dimensions["L"].width=20
    sheet.column_dimensions["M"].width=20
    
    sheet.row_dimensions[1].height = 45
    
    book.save(report_path)
     





def create_excel_by_local_out_of_balance(report_path, report_type, locals, db_session):
    book=openpyxl.Workbook()
    sheet = book.active
    start_timer = time.time()
    get_by_copy = db_session.execute(text(f''' 
                                        SELECT 
                                            CR.NAME AS REGION,
                                            LC.CODE AS LOCAL_CODE,
                                            FT.ID AS FROM_TYPE_ID,
                                            SUM(CASE WHEN BT.ACCOUNT_95413_START_STATE IS NOT NULL THEN BT.ACCOUNT_95413_START_STATE::real ELSE 0 END) as ACCOUNT_95413_START_STATE,
                                            SUM(CASE WHEN BT.ACCOUNT_95413_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_95413_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_95413_DEBIT_SUM,
                                            SUM(CASE WHEN BT.ACCOUNT_95413_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_95413_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_95413_CREDIT_SUM,
                                            SUM((
                                            CASE WHEN BT.ACCOUNT_95413_START_STATE IS NOT NULL THEN BT.ACCOUNT_95413_START_STATE::real ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_95413_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_95413_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_95413_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_95413_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_95413_BALANCE,

                                            SUM(CASE WHEN BT.ACCOUNT_9150X_START_STATE IS NOT NULL THEN (BT.ACCOUNT_9150X_START_STATE::real) ELSE 0 END) as ACCOUNT_9150X_START_STATE,
                                            SUM(CASE WHEN BT.ACCOUNT_9150X_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_9150X_DEBIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_9150X_DEBIT_SUM,
                                            SUM(CASE WHEN BT.ACCOUNT_9150X_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_9150X_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_9150X_CREDIT_SUM,
                                            SUM((CASE WHEN BT.ACCOUNT_9150X_START_STATE IS NOT NULL THEN BT.ACCOUNT_9150X_START_STATE::real ELSE 0 END
                                            +
                                            CASE WHEN BT.ACCOUNT_9150X_DEBIT_SUM IS NOT NULL THEN (BT.ACCOUNT_9150X_DEBIT_SUM::real)/100 ELSE 0 END)
                                            -
                                            CASE WHEN BT.ACCOUNT_9150X_CREDIT_SUM IS NOT NULL THEN (BT.ACCOUNT_9150X_CREDIT_SUM::real)/100 ELSE 0 END) as ACCOUNT_9150X_BALANCE
                                        FROM BALANCE_TURNOVER BT
                                        JOIN LOAN_PORTFOLIO LP ON BT.LOAN_ID = LP.LOAN_ID
                                        JOIN CLIENT_REGION CR ON LP.CLIENT_REGION = CR.ID
                                        JOIN LOCAL_CODE LC ON LP.LOCAL_CODE_ID = LC.ID
                                        {report_type}
                                        AND LP.LOCAL_CODE_ID {locals}
                                        GROUP BY LC.CODE, CR.NAME, FT.ID
                                          ''')).fetchall()

 
    print(time.time() - start_timer)
    row = 2
    i=1
    for loan in get_by_copy:
        i=i+1
        
        sheet.cell(column=1, row=row).value = i
        sheet.cell(column=2, row=row).value = loan['region']
        sheet.cell(column=3, row=row).value = loan['local_code']
        sheet.cell(column=4, row=row).value = loan['account_95413_start_state']
        sheet.cell(column=5, row=row).value = loan['account_95413_debit_sum']
        sheet.cell(column=6, row=row).value = loan['account_95413_credit_sum']
        sheet.cell(column=7, row=row).value = loan['account_95413_balance']
        sheet.cell(column=8, row=row).value = loan['account_9150x_start_state']
        sheet.cell(column=9, row=row).value = loan['account_9150x_debit_sum']
        sheet.cell(column=10, row=row).value = loan['account_9150x_credit_sum']
        sheet.cell(column=11, row=row).value = loan['account_9150x_balance']
        
        row = row + 1
    
    #sheet.insert_rows(1)
    sheet.cell(column=1, row=1).value = "ID"
    sheet.cell(column=2, row=1).value =  "Регион"
    sheet.cell(column=3, row=1).value = "Локал код"
    sheet.cell(column=4, row=1).value = "Остаток 95413 на начало отчетного периода"
    sheet.cell(column=5, row=1).value = "Зачислено в течении месяца по счёту 95413"
    sheet.cell(column=6, row=1).value = "Погашено за отчетный период по счёту 95413"
    sheet.cell(column=7, row=1).value = "Остаток 95413 к концу отчетного периода"
    sheet.cell(column=8, row=1).value = "Остаток 91501 и 91503 на начало отчетного периода"
    sheet.cell(column=9, row=1).value = "Зачислено в течении месяца по счетам 91501 и 91503"
    sheet.cell(column=10, row=1).value = "Погашено за отчетный период по счетам 91501 и 91503"
    sheet.cell(column=11, row=1).value = "Остаток 91501 и 91503 к концу отчетного периода"
    
    bold = Font(bold=True)

    for cell in sheet["1:1"]:
        cell.font = bold
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    sheet.column_dimensions["A"].width=8
    sheet.column_dimensions["B"].width=8
    sheet.column_dimensions["C"].width=8
    sheet.column_dimensions["D"].width=20
    sheet.column_dimensions["E"].width=20
    sheet.column_dimensions["F"].width=20
    sheet.column_dimensions["G"].width=20
    sheet.column_dimensions["H"].width=20
    sheet.column_dimensions["I"].width=20
    sheet.column_dimensions["J"].width=20
    sheet.column_dimensions["K"].width=20
    
    sheet.row_dimensions[1].height = 45
    
    book.save(report_path)
    print('end: ', time.time() - start_timer, " sec")