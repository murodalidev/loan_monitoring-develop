import openpyxl

from openpyxl.styles import Font, Alignment
from fastapi.responses import FileResponse
import datetime
from fastapi import HTTPException

from ...models.users.users import Users

from ...models.users.attached_branches import attached_branches
from ..loan_monitoring.report_order import report_order_crud
from ...common.dictionaries.general_tasks_dictionary import  MGT
from ...common.dictionaries.departments_dictionary import ROLES
from ...common.commit import commit_object, flush_object
from ...config.logs_config import info_logger
import uuid, json
from ...common.decorator import measure_time

import time
from ..loan_monitoring.integrations.soliq_integrations import get_eauksion_orders

@measure_time
def create_report_to_excel(request, user_id, db_session):
    report_path = f"project_files/loan_reports/eauksion_{str(datetime.datetime.now().date())}_{str(uuid.uuid4())[:4]}.xlsx"
    report_order_crud.report_order(user_id, report_path, request.report_type, request.report_by, db_session)
    
    create_excel(report_path, request.request, user_id, db_session)
        
    report_order_crud.change_report_status_to_ready(user_id, request.report_type, request.report_by, db_session)
    commit_object(db_session)
    print('report ready')
    info_logger.info("User %s requested report for loan monitoring", user_id)
    return FileResponse(report_path, filename=f"{'auksion_'+str(datetime.datetime.now().date())}.xlsx")
    




def create_excel(report_path, request, user_id, db_session):
    book=openpyxl.Workbook()
    sheet = book.active
    start_timer = time.time()
    get_data =  get_eauksion_orders(request, user_id, db_session)
    row = 2
    i=1
    for data in get_data['orders']:
        
        sheet.cell(column=1, row=row).value = i
        sheet.cell(column=2, row=row).value = data['lot']
        sheet.cell(column=3, row=row).value = data['property_name']
        sheet.cell(column=4, row=row).value = data['property_category']
        sheet.cell(column=5, row=row).value = data['property_group']
        sheet.cell(column=6, row=row).value = data['status']
        sheet.cell(column=7, row=row).value = data['price']
        sheet.cell(column=8, row=row).value = data['start_price']
        sheet.cell(column=9, row=row).value = data['sold_price']
        sheet.cell(column=10, row=row).value = data['exec_doc_date']
        sheet.cell(column=11, row=row).value = data['exec_doc_number']
        
        row = row + 1
        i = i + 1
    
    #sheet.insert_rows(1)
    
    sheet.cell(column=1, row=1).value = "№"
    sheet.cell(column=2, row=1).value = "Лот"
    sheet.cell(column=3, row=1).value = "Наименование имущества"
    sheet.cell(column=4, row=1).value = "Категория имущества"
    sheet.cell(column=5, row=1).value = "Группа объектов имущества"
    sheet.cell(column=6, row=1).value = "Статус"
    sheet.cell(column=7, row=1).value = "Оценосная стоимость"
    sheet.cell(column=8, row=1).value = "Начальная стоимость"
    sheet.cell(column=9, row=1).value = "Цена продажи"
    sheet.cell(column=10, row=1).value = "Дата аукционы"
    sheet.cell(column=11, row=1).value = "Номер испонительного документа"
    
    bold = Font(bold=True)

    # for cell in sheet["1:1"]:
    #     cell.font = bold
    # sheet.row_dimensions[1].height = 30
    # sheet.merge_cells('I1:M1')
    # megre_cell = sheet['I1']
    # megre_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    
    # sheet.merge_cells('N1:R1')
    # megre_cell = sheet['N1']
    # megre_cell.alignment = Alignment(horizontal="center", vertical="center")
    # sheet.merge_cells('S1:W1')
    # megre_cell = sheet['S1']
    # megre_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    
    for i in range(11):
        sheet[1][i].font = bold
        sheet[1][i].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    sheet.column_dimensions["A"].width=8
    sheet.column_dimensions["B"].width=12
    sheet.column_dimensions["C"].width=50
    sheet.column_dimensions["D"].width=40
    sheet.column_dimensions["E"].width=16
    sheet.column_dimensions["F"].width=30
    sheet.column_dimensions["G"].width=12
    sheet.column_dimensions["H"].width=12
    sheet.column_dimensions["I"].width=12
    sheet.column_dimensions["J"].width=12
    sheet.column_dimensions["K"].width=23
    
    sheet.row_dimensions[1].height = 30
    
    book.save(report_path)
    print('end: ', time.time() - start_timer, " sec")