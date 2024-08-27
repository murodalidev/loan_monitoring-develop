import openpyxl

from openpyxl.styles import Font, Alignment
import datetime
from fastapi.responses import FileResponse
import json
from app.services.loan_monitoring.integrations import adliya_integrations
from .report_to_excel_service import check_user_acces_to_locals
from ..loan_monitoring.report_order import report_order_crud
from ...models.statuses.from_type_model import FromType
from ...common.dictionaries.general_tasks_dictionary import  MGT
from ...common.commit import commit_object, flush_object
from ...common.decorator import measure_time
from ...config.logs_config import info_logger
import uuid, json
import time
from sqlalchemy.sql import text



@measure_time
def create_adliya_report_to_excel(request, user_id, db_session):
    #response = check_user_acces_to_locals(user_id, request.report_type, db_session)
    
    
    
    report_path = f"project_files/loan_reports/adliya_{str(datetime.datetime.now().date())}_{str(uuid.uuid4())[:4]}.xlsx"
    report_order_crud.report_order(user_id, report_path, request.report_type, 1, db_session)
    
    create_excel_for_adliya(request.start_date, request.end_date, report_path, db_session)
        
        
    report_order_crud.change_report_status_to_ready(user_id, request.report_type, 1, db_session)
    commit_object(db_session)
    info_logger.info("User %s requested report for adliya", user_id)
    return FileResponse(report_path, filename=f"adliya'_'+{str(datetime.datetime.now().date())}.xlsx")








def create_excel_for_adliya(start_date, end_date, report_path, db_session):
    
    start_date_filter =''
    end_date_filter = ''
    if start_date is not None:
        start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
        start_date_filter = f"and created_at::date >='{start_date}'::date"
        
    if end_date is not None:
        end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        end_date_filter = f"and created_at::date <='{end_date}'::date"
    
    org_type = adliya_integrations.adliya_org_type(3, db_session)
    
    orgs= {}
    for org in org_type['result']:
        orgs[org['id']] = org['nameuz']
    book=openpyxl.Workbook()
    sheet = book.active
    get_by_copy = db_session.execute(text(f''' 
                                          select * from integrations_log where service_api_id in (4,5)
                                          {start_date_filter}
                                          {end_date_filter}
                                          order by id asc
                                        
                                          ''')).fetchall()

    row = 2
    i=0
    
    for loan in get_by_copy:
        i=i+1
        
        sheet.cell(column=1, row=row).value = i
        sheet.cell(column=2, row=row).value = loan['created_at']
        sheet.cell(column=3, row=row).value = json.loads(loan[4])['params']['reg_num']
        if json.loads(loan[4])['params']['statement']['doc_type'] == '39':
            sheet.cell(column=4, row=row).value = 'Письмо Банка'
        if json.loads(loan[4])['params']['statement']['doc_type'] == '1':
            sheet.cell(column=4, row=row).value = 'Судебный акт'
        if json.loads(loan[4])['params']['statement']['doc_type'] == '2':
            sheet.cell(column=4, row=row).value = 'Постановление суда'
        if json.loads(loan[4])['params']['statement']['doc_type'] == '31':
            sheet.cell(column=4, row=row).value = 'Другие случаи, предусмотренные законодательством'
        sheet.cell(column=5, row=row).value = json.loads(loan[4])['params']['statement']['doc_num']
        sheet.cell(column=6, row=row).value = json.loads(loan[4])['params']['statement']['doc_date']
        sheet.cell(column=7, row=row).value = json.loads(loan[4])['params']['statement']['org_name']
        sheet.cell(column=8, row=row).value = json.loads(loan[4])['params']['statement']['org_fio']
        sheet.cell(column=9, row=row).value = json.loads(loan[4])['params']['declarant']['company_name'] 
        if 'result' in loan[5]:
            sheet.cell(column=10, row=row).value = loan[5]['result']['res_note']
        
            sheet.cell(column=11, row=row).value = loan[5]['result']['reg_date']
        
            sheet.cell(column=12, row=row).value = orgs[loan[5]['result']['ban_info']['org_type']]
        
            sheet.cell(column=13, row=row).value = loan[5]['result']['ban_info']['org_fio']
        
            sheet.cell(column=14, row=row).value = loan[5]['result']['subject']['district']
        
            sheet.cell(column=15, row=row).value = loan[5]['result']['subject']['street']
            home = ''
        
            if 'home' in loan[5]['result']['subject']:
                home = loan[5]['result']['subject']['home']
            sheet.cell(column=16, row=row).value = home
            
        
            sheet.cell(column=17, row=row).value = loan[5]['result']['subject']['state_num']
        
            sheet.cell(column=18, row=row).value = loan[5]['result']['subject']['engine_num']
            sheet.cell(column=19, row=row).value = loan[5]['result']['subject']['body_num']
            sheet.cell(column=20, row=row).value = loan[5]['result']['subject']['chassis_num']
            sheet.cell(column=21, row=row).value = loan[5]['result']['subject']['mark']
            sheet.cell(column=22, row=row).value = loan[5]['result']['subject']['year_create']
            sheet.cell(column=23, row=row).value = loan[5]['result']['subject']['color']
            sheet.cell(column=24, row=row).value = loan[5]['result']['subject']['tech_serial']
            sheet.cell(column=25, row=row).value = loan[5]['result']['subject']['tech_num']
            sheet.cell(column=26, row=row).value = loan[5]['result']['subject']['tech_date']
            sheet.cell(column=27, row=row).value = loan[5]['result']['subject']['tech_issue']
        elif 'error' in loan[5]:
            if 'data' in loan[5]['error']:
                sheet.cell(column=10, row=row).value = loan[5]['error']['data']
            else:
                sheet.cell(column=10, row=row).value = loan[5]['error']['message']
        row = row + 1
    
    
    sheet.insert_rows(1)
    sheet.cell(column=1, row=1).value = "№"
    sheet.cell(column=2, row=1).value =  "Дата и время отправки"
    sheet.cell(column=3, row=1).value = "Регистрационный номер запрета"
    sheet.cell(column=4, row=1).value = "Тип документа основания снятия ареста"
    sheet.cell(column=5, row=1).value = "Номер документа основания снятия ареста"
    sheet.cell(column=6, row=1).value = "Дата документа основания снятия ареста"
    sheet.cell(column=7, row=1).value = "Наименование организации, снявший запрет/арест"
    sheet.cell(column=8, row=1).value = "ФИО лица организации, снявшего запрета"
    sheet.cell(column=9, row=1).value = "Наименование компании (Клиент)" 
    sheet.cell(column=10, row=1).value = "Ответ снятия ареста"
    sheet.cell(column=11, row=1).value = "Дата снятия ареста"
    sheet.cell(column=12, row=1).value = "Информация о запрете (Орган)"
    sheet.cell(column=13, row=1).value = "Информация о запрете (ФИО нотариуса)"
    sheet.cell(column=14, row=1).value = "Район"
    sheet.cell(column=15, row=1).value = "Улица"
    sheet.cell(column=16, row=1).value = "Дом"
    sheet.cell(column=17, row=1).value = "Гос.номер автотранспорта"
    sheet.cell(column=18, row=1).value = "Серийный номер двигателя"
    sheet.cell(column=19, row=1).value = "Серийный номер кузова"
    sheet.cell(column=20, row=1).value = "Серийный номер шоссе"
    sheet.cell(column=21, row=1).value = "Марка"
    sheet.cell(column=22, row=1).value = "Дата производства"
    sheet.cell(column=23, row=1).value = "Цвет"
    sheet.cell(column=24, row=1).value = "Серия тех. паспорта"
    sheet.cell(column=25, row=1).value = "Номер тех. паспорта"
    sheet.cell(column=26, row=1).value = "Дата выдачи тех. паспорта"
    sheet.cell(column=27, row=1).value = "Кем выдан тех. паспорта"
    
    bold = Font(bold=True)

    for row in sheet["1:2"]:
        for cell in row:
            cell.font = bold
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    
    sheet.column_dimensions["A"].width=10
    sheet.column_dimensions["B"].width=25
    sheet.column_dimensions["C"].width=25
    sheet.column_dimensions["D"].width=25
    sheet.column_dimensions["E"].width=25
    sheet.column_dimensions["F"].width=25
    sheet.column_dimensions["G"].width=25
    sheet.column_dimensions["H"].width=25
    sheet.column_dimensions["I"].width=25
    sheet.column_dimensions["J"].width=25
    sheet.column_dimensions["K"].width=25
    sheet.column_dimensions["L"].width=25
    sheet.column_dimensions["M"].width=25
    sheet.column_dimensions["N"].width=25
    sheet.column_dimensions["O"].width=25
    sheet.column_dimensions["P"].width=25
    sheet.column_dimensions["Q"].width=25
    sheet.column_dimensions["R"].width=25
    sheet.column_dimensions["S"].width=25
    sheet.column_dimensions["T"].width=25
    sheet.column_dimensions["U"].width=25
    sheet.column_dimensions["V"].width=25
    sheet.column_dimensions["W"].width=25
    sheet.column_dimensions["X"].width=25
    sheet.column_dimensions["Y"].width=25
    sheet.column_dimensions["Z"].width=25
    sheet.column_dimensions["AA"].width=25
    
    book.save(report_path)