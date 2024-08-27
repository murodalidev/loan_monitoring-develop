from openpyxl import load_workbook
import time, datetime
from datetime import timedelta
from ....models.brief_case.loan_portfolio import Loan_Portfolio
from ....models.brief_case.directories.local_code import local_code
from ..directories.load_all_from_files import get_directories
from .get_digit_from_str import digit_from_str
from ....config.logs_config import cron_logger
from ....common.commit import commit_object

def load_portfolio(db_session):
    start_timer = time.time()
    yesterday = (datetime.datetime.now()-timedelta(days=1)).date().strftime("%d-%m-%Y")
    
    cron_logger.info(f'Started loading portfolio, start_time: {start_timer}')
    #date = str(datetime.datetime.now().date().strftime('%m_%d_%Y'))
    directories = get_directories(db_session)
    path_to_portfolio = f'project_files/loan_portfolio/portfolio_{yesterday}.xlsx'
    i=0
    try:
        book = load_workbook(path_to_portfolio, data_only=True,  read_only = True, keep_vba = False)
    except FileNotFoundError  as e:
        cron_logger.error(e)
        return 0
    sheet = book.active
    rows = sheet.rows
    start = False
    strokes = next(rows)
    for _ in range(0, sheet.max_row):
        if strokes[0].value == 'NN':
            start = True
            continue
        if start:
            if strokes[0].value != None:
                i=i+1
                if i%10000==0:
                    cron_logger.info(f'added {i} elements to portfolio')
                    
                local = db_session.query(local_code).filter(local_code.code == strokes[4].value).first()
                if local is None:
                    
                    new_local = local_code(code = str(strokes[4].value))
                    db_session.add(new_local)
                    commit_object(db_session)
                    directories = get_directories(db_session)
                    cron_logger.info(f'detected new local code:  {strokes[4].value}')
                create_portfolio_data(strokes, db_session, directories)
       
     
    commit_object(db_session)
    end_timer = time.time()
    res = end_timer - start_timer
    final_res = res / 60
    cron_logger.info(f'Finished loading portfolio data')
    cron_logger.info(f'Execution time: {final_res} seconds')
    return {"result":"OK"}




def create_portfolio_data(strokes, db_session, directories=None):
    portfolio = Loan_Portfolio(
                    client_region =strokes[1].value != None and directories['client_region'][strokes[1].value] or None,
                    client_district =strokes[2].value,
                    local_code_id = directories['local_code'][strokes[5].value],
                    loan_id = strokes[6].value,
                    grki_contract=strokes[7].value,
                    client_name=strokes[8].value,
                    balance_account=strokes[9].value,
                    loan_account=strokes[10].value,
                    num_and_date_decision=strokes[11].value,
                    currency_id=directories['currencies'][strokes[12].value],
                    contract_amount=strokes[13].value,
                    contract_amount_uz_currency=strokes[14].value,
                    issue_date=strokes[15].value.date(),
                    maturity_date=strokes[16].value.date(),
                    term_type=strokes[17].value,
                    num_and_date_contract=strokes[18].value,
                    osn_cmp_percent=strokes[19].value,
                    # loan_purpose=strokes[20].value,  # New column added
                    # loan_source=strokes[21].value,   # New column added
                    overdue_percentage=strokes[22].value,
                    credit_account_balance=strokes[23].value,
                    balance_rev=strokes[24].value,
                    num_and_date_renewal=strokes[25].value,
                    maturity_date_renewal=strokes[26].value and strokes[26].value.date() or None,
                    overdue_balance=strokes[27].value,
                    overdue_start_date=strokes[28].value and strokes[28].value.date() or None,
                    judicial_balance=strokes[29].value,
                    law_enforcement_code=strokes[30].value,
                    solution_sign=strokes[31].value,
                    date_transf_date_decision=strokes[32].value,
                    total_overdue=strokes[33].value,
                    quality_class=strokes[34].value,
                    reserve_balance=strokes[35].value,
                    balance_init_percent=strokes[36].value,
                    balance_init_overdue_percent=strokes[37].value,
                    security_valuation=strokes[38].value,
                    loan_security=strokes[39].value,
                    security_description=strokes[40].value,
                    funds_sources=strokes[41].value,
                    lending_type=strokes[42].value,
                    loan_goal=strokes[43].value,
                    client_parent_org=strokes[44].value,
                    loan_industry=strokes[45].value,
                    client_industry=strokes[46].value,
                    credit_rating=strokes[47].value,
                    cb_chairman=strokes[48].value,
                    client_address=strokes[49].value,
                    contract_uid=strokes[50].value,
                    inn_passport=strokes[51].value,
                    balance_interest_offbalance=strokes[52].value,
                    overdue_balance_interest_offbalance=strokes[53].value,
                    sum_cur_year=strokes[54].value,
                    repaid_cur_year=strokes[55].value,
                    total_issued=strokes[56].value,
                    total_repaid=strokes[57].value,
                    loan_purpose_specific=strokes[58].value,
                    credit_line_purpose=strokes[59].value,
                    judicial_account=strokes[60].value,
                    judicial_account_writeoff=strokes[61].value,
                    debt_account=strokes[62].value,
                    inps=strokes[63].value,
                    birth_date=strokes[64].value and strokes[64].value.date() or None,
                    balance_16309=strokes[65].value,
                    balance_16379=strokes[66].value,
                    balance_16323=strokes[67].value,
                    balance_16325=strokes[68].value,
                    balance_16377=strokes[69].value,
                    date_overdue_percent=strokes[70].value,
                    balance_16397=strokes[71].value,
                    balance_91501=strokes[72].value,
                    balance_91503=strokes[73].value,
                    balance_95413=strokes[74].value,
                    balance_XXX99=strokes[75].value,
                    balance_91809=strokes[76].value,
                    credit_status=strokes[77].value,
                    checking_account_balance=strokes[78].value,
                    loan_product=strokes[79].value and str(strokes[79].value).replace(" ", "") or None,
                    borrower_type=strokes[81].value,
                    num_workplaces=strokes[82].value,
                    gender_id=strokes[83].value and directories['genders'][digit_from_str(strokes[83].value)] or None,
                    status = directories['loan_status']['01'],
                    checked_status = 1
                    )
    db_session.add(portfolio)