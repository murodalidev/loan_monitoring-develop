import datetime
from openpyxl import load_workbook
import time
import os
from datetime import timedelta
from fastapi import HTTPException
from sqlalchemy.sql import text
from ....models.brief_case.loan_portfolio import Loan_Portfolio, Loan_Portfolio_Date
from ....models.brief_case.directories.local_code import local_code
from ..directories.load_all_from_files import get_directories
from .get_digit_from_str import digit_from_str
from ....common.commit import commit_object
from ....common.decorator import measure_time
from .loan_portfolio_service import create_portfolio_data
from ....config.logs_config import cron_logger

def portfolio_update_full(date, db_session):
    # get_last = db_session.query(local_code).order_by(local_code.id.desc()).first()
    # db_session.execute(text(f'ALTER SEQUENCE local_code_id_seq RESTART WITH {get_last.id+1}'))
    # db_session.commit()
    # exit()
    cron_logger.info('started uncheck portfolio script')
    uncheck_portfolio_data(db_session)
    cron_logger.info('finished uncheck portfolio script')
    cron_logger.info('started update portfolio script')
    update_portfolio(date, db_session)
    cron_logger.info('finished update portfolio script')
    cron_logger.info('started close portfolio script')
    closed_portfolio_data(db_session)
    cron_logger.info('finished close portfolio script')
    return "OK"



@measure_time
def update_portfolio(date, db_session):
    res = os.system(f"scp user@172.23.7.2:/home/user/repositories/loan_monitoring/backend/project_files/loan_portfolio/{date}.xlsx /home/israilovk/repositories/loan_monitoring/project_files/loan_portfolio")
    
    
    
    if date is None:
        current_date = (datetime.datetime.now()-timedelta(days=1)).date().strftime("%d-%m-%Y")
    else:
        current_date = datetime.datetime.strptime(date.replace('.','-'), "%d-%m-%Y").date()
    get_portfolio_date = db_session.query(Loan_Portfolio_Date).filter(Loan_Portfolio_Date.date == current_date).first()
    
    if get_portfolio_date is None:
        directories = get_directories(db_session)
        path_to_portfolio = f'project_files/loan_portfolio/{date}.xlsx'
        i=0
        try:
            book = load_workbook(path_to_portfolio, data_only=True, read_only = True, keep_vba = False)
        except FileNotFoundError  as e:
            cron_logger.error(e)
            raise HTTPException(status_code=403, detail=str(e))
        sheet = book.active
        rows = sheet.rows
        start = False
        for _ in range(0, sheet.max_row):
                
            strokes = next(rows)
            if strokes[0].value == 'NN':
                start = True
                continue
            if start:
                
                if strokes[0].value != None:
                    local = db_session.query(local_code).filter(local_code.code == strokes[5].value).first()
                    if local is None:
                        print(strokes[5].value, directories['client_region'][digit_from_str(strokes[1].value)])
                        new_local = local_code(code = str(strokes[5].value), region_id = directories['client_region'][digit_from_str(strokes[1].value)], status = False)
                        db_session.add(new_local)
                        commit_object(db_session)
                        directories = get_directories(db_session)
                    portfolio = db_session.query(Loan_Portfolio).filter(Loan_Portfolio.loan_id == strokes[6].value).first()
                    
                    i=i+1
                    if i%10000==0:
                        cron_logger.info(f'Updated {i} protfolio elemnts')
                    if portfolio is not None:
                        if strokes[1].value is not None:
                            portfolio.client_region = directories['client_region'][digit_from_str(strokes[1].value)]
                        portfolio.client_district = strokes[2].value
                        portfolio.local_code_id = directories['local_code'][digit_from_str(strokes[5].value)]
                        portfolio.loan_id = strokes[6].value
                        portfolio.grki_contract = strokes[7].value
                        portfolio.client_name = strokes[8].value
                        portfolio.balance_account = strokes[9].value
                        portfolio.loan_account = strokes[10].value
                        portfolio.num_and_date_decision = strokes[11].value
                        portfolio.currency_id = directories['currencies'][strokes[12].value]
                        portfolio.contract_amount = strokes[13].value
                        portfolio.contract_amount_uz_currency = strokes[14].value
                        portfolio.issue_date = strokes[15].value.date()
                        portfolio.maturity_date = strokes[16].value.date()
                        portfolio.term_type = strokes[17].value
                        portfolio.num_and_date_contract = strokes[18].value
                        portfolio.osn_cmp_percent = strokes[19].value
                        # portfolio.loan_purpose = strokes[20].value  # New column added
                        # portfolio.loan_source = strokes[21].value   # New column added
                        portfolio.overdue_percentage = strokes[22].value
                        portfolio.credit_account_balance = strokes[23].value
                        portfolio.balance_rev = strokes[24].value
                        portfolio.num_and_date_renewal = strokes[25].value
                        portfolio.maturity_date_renewal = strokes[26].value and strokes[26].value.date() or None
                        portfolio.overdue_balance = strokes[27].value
                        portfolio.overdue_start_date = strokes[28].value and strokes[28].value.date() or None
                        portfolio.judicial_balance = strokes[29].value
                        portfolio.law_enforcement_code = strokes[30].value
                        portfolio.solution_sign = strokes[31].value
                        portfolio.date_transf_date_decision = strokes[32].value
                        portfolio.total_overdue = strokes[33].value
                        portfolio.quality_class = strokes[34].value
                        portfolio.reserve_balance = strokes[35].value
                        portfolio.balance_init_percent = strokes[36].value
                        portfolio.balance_init_overdue_percent = strokes[37].value
                        portfolio.security_valuation = strokes[38].value
                        portfolio.loan_security = strokes[39].value
                        portfolio.security_description = strokes[40].value
                        portfolio.funds_sources = strokes[41].value
                        portfolio.lending_type = strokes[42].value
                        portfolio.loan_goal = strokes[43].value
                        portfolio.client_parent_org = strokes[44].value
                        portfolio.loan_industry = strokes[45].value
                        portfolio.client_industry = strokes[46].value
                        portfolio.credit_rating = strokes[47].value
                        portfolio.cb_chairman = strokes[48].value
                        portfolio.client_address = strokes[49].value
                        portfolio.contract_uid = strokes[50].value
                        portfolio.inn_passport = strokes[51].value
                        portfolio.balance_interest_offbalance = strokes[52].value
                        portfolio.overdue_balance_interest_offbalance = strokes[53].value
                        portfolio.sum_cur_year = strokes[54].value
                        portfolio.repaid_cur_year = strokes[55].value
                        portfolio.total_issued = strokes[56].value
                        portfolio.total_repaid = strokes[57].value
                        portfolio.loan_purpose_specific = strokes[58].value
                        portfolio.credit_line_purpose = strokes[59].value
                        portfolio.judicial_account = strokes[60].value
                        portfolio.judicial_account_writeoff = strokes[61].value
                        portfolio.debt_account = strokes[62].value
                        portfolio.inps = strokes[63].value
                        try:
                            strokes[64].value and strokes[64].value.date() or None
                        except:
                            print(strokes[64].value)
                        portfolio.birth_date = strokes[64].value and strokes[64].value.date() or None
                        portfolio.balance_16309 = strokes[65].value
                        portfolio.balance_16379 = strokes[66].value
                        portfolio.balance_16323 = strokes[67].value
                        portfolio.balance_16325 = strokes[68].value
                        portfolio.balance_16377 = strokes[69].value
                        portfolio.date_overdue_percent = strokes[70].value
                        portfolio.balance_16397 = strokes[71].value
                        portfolio.balance_91501 = strokes[72].value
                        portfolio.balance_91503 = strokes[73].value
                        portfolio.balance_95413 = strokes[74].value
                        portfolio.balance_XXX99 = strokes[75].value
                        portfolio.balance_91809 = strokes[76].value
                        portfolio.credit_status = strokes[77].value
                        portfolio.checking_account_balance = strokes[78].value
                        portfolio.loan_product = strokes[79].value and str(strokes[79].value).replace(" ", "") or None
                        portfolio.borrower_type = strokes[81].value
                        portfolio.num_workplaces = strokes[82].value
                        portfolio.gender_id = strokes[83].value and directories['genders'][digit_from_str(strokes[83].value)] or None
                        portfolio.updated = datetime.datetime.now()
                        portfolio.checked_status = 1
                        db_session.flush()
                    else:
                        create_portfolio_data(strokes, db_session, directories)
                    #db_session.add(portfolio)
        new_portfolio_date = Loan_Portfolio_Date(date = current_date, created = datetime.datetime.now())
        db_session.add(new_portfolio_date)
    try:
        db_session.commit()
    except:
        db_session.rollback()
        raise  
    return {"result":"OK"}


        


@measure_time
def uncheck_portfolio_data(db_session):
    db_session.execute( '''UPDATE LOAN_PORTFOLIO SET CHECKED_STATUS=2, STATUS=1''' )
    commit_object(db_session)
    
    
@measure_time
def closed_portfolio_data(db_session):
    
    db_session.execute( '''UPDATE LOAN_PORTFOLIO SET STATUS=3, UPDATED=NOW() WHERE CHECKED_STATUS=2''' )
    commit_object(db_session)
    
    
    
    


def get_portfolio_data(db_session):
    return db_session.query(Loan_Portfolio_Date).order_by(Loan_Portfolio_Date.date.desc()).first()