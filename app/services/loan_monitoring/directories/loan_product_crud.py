import datetime
from app.models.brief_case.directories.loan_product import loan_product, loan_product_edit
from app.models.brief_case.loan_portfolio import Loan_Portfolio
from app.models.brief_case.directories.loan_product_type import loan_product_type
from openpyxl import load_workbook
import time
from ....common.is_empty import is_empty, is_exists
from ....common.commit import commit_object
from ....config.logs_config import info_logger

def load_loan_product(db_session):
    
    book = load_workbook('project_files/directories/loan_product.xlsx',  read_only = True, keep_vba = False)
    sheet = book.active
    rows = sheet.rows
    start = False
    for i in range(0, sheet.max_row):
            
        strokes = next(rows)
        if strokes[0].value == 'Кредитный продукт':
            
            start = True
            continue
        if start:
            
            if strokes[0].value != None:
                product = loan_product(name = strokes[1].value, is_target = strokes[2].value, type = strokes[3].value)
                db_session.add(product)
    try:
        db_session.commit()
    except:
        db_session.rollback()
        raise
    return {"result":"OK"}






def create_loan_product(request, username, db_session):
    username = username
    info_logger.info(f"user {username} called create_loan_product api")
    info_logger.info(f"user {username} created {request.loan_product_name} product")
    get_dep = db_session.query(loan_product).filter(loan_product.name == request.loan_product_name).first()
    is_empty(get_dep, 400, 'loan product has already created!')
    
    new_loan_product = loan_product(name=request.loan_product_name,
                                    type=request.loan_product_type_id,
                                    is_target = request.is_target,
                                    checked = True,
                                    created_at = datetime.datetime.now())
    db_session.add(new_loan_product)
    commit_object(db_session)
            
            
    
    return {"result":"OK"}



def get_all_loan_products(name, is_target, type, checked, page, size, db_session):
    get_products = db_session.query(loan_product).order_by(loan_product.id.asc())
    
    if name is not None:
        get_products = get_products.filter(loan_product.name.like(f"%{name}%"))
    if is_target is not None:
        if is_target == 2:
            is_target = None
        get_products = get_products.filter(loan_product.is_target == is_target)
    if type is not None:
        if type == 0:
            type = None
        get_products = get_products.filter(loan_product.type == type)
    if checked is not None:
        get_products = get_products.filter(loan_product.checked == checked)
    count = get_products.count()
    get_products = get_products.limit(size).offset((page-1)*size).all()
    
    loan_products = []
    for prod in get_products:
        loan_products.append({"id":prod.id,
               "name":prod.name,
               "type":prod.type and {"id":prod.loan_type.id,
                                     "name":prod.loan_type.name} or None,
               "is_target":prod.is_target})
        
        
    return {"items":loan_products,
            "total":count,
            "page":page,
            "size":size}




def get_loan_product_type(db_session):
    get_type = db_session.query(loan_product_type).all()
    product_type = []
    for type in get_type:
        product_type.append({"id":type.id,
               "name":type.name})
    return product_type




def get_loan_product(loan_product_id, db_session):
    return db_session.query(loan_product).filter(loan_product.id == loan_product_id).first()



def update_loan_product(id, request, username, db_session):
    username = username
    info_logger.info(f"user {username} called update_loan_product api")
    get_loan_product = db_session.query(loan_product).filter(loan_product.id == id).first()
    is_exists(get_loan_product, 400, 'loan product doesn`t exist!')
    
    name = None
    product_type = None
    is_target = None
    
    if request.loan_product_name:
        info_logger.info(f"user {username} edited name from:{get_loan_product.name} to: {request.loan_product_name}")
        name = get_loan_product.name
        get_loan_product.name=request.loan_product_name
    if request.loan_product_type_id:
        info_logger.info(f"user {username} edited type from:{get_loan_product.type} to: {request.loan_product_type_id}")
        product_type = get_loan_product.type
        get_loan_product.type = request.loan_product_type_id
    if request.is_target is not None:
        info_logger.info(f"user {username} edited is_target from:{get_loan_product.is_target} to: {request.is_target}")
        is_target = get_loan_product.is_target
        get_loan_product.is_target= request.is_target
    if (request.loan_product_type_id is not None or get_loan_product.type is not None) and (request.is_target is not None or get_loan_product.is_target is not None):
        get_loan_product.checked = True
    get_loan_product.updated_at = datetime.datetime.now()
    
    
    if name or product_type or is_target is not None:
        
        edit_product = loan_product_edit(
            name_old = name,
            name_new = request.loan_product_name,
            type_old = product_type,
            type_new = request.loan_product_type_id,
            is_target_old = is_target,
            is_target_new = request.is_target,
            created_at = datetime.datetime.now()
        )
        db_session.add(edit_product)
    
    commit_object(db_session)
    
    # if request.is_target is not None:
    return {"result":"OK"}



def delete_loan_product(id, username, db_session):
    info_logger.info(f"user {username} called delete_loan_product api")
    
    get_dep = db_session.query(loan_product).filter(loan_product.id == id).first()
    is_exists(get_dep, 400, 'Loan product not found')
    db_session.delete(get_dep)
    commit_object(db_session)
    info_logger.info(f"user {username} deleted product {id}")
    return {"result":"OK"}


def update_loan_product_table(db_session):
    get_all_loan_product = db_session.query(Loan_Portfolio.loan_product).distinct(Loan_Portfolio.loan_product).all()
    
    for product in get_all_loan_product:
        if product.loan_product:
            get_product = db_session.query(loan_product).filter(loan_product.name == product.loan_product).first()
            if get_product is None:
                new_loan_product = loan_product(name=product.loan_product)
                db_session.add(new_loan_product)
    commit_object(db_session)
    return {"result":"OK"}