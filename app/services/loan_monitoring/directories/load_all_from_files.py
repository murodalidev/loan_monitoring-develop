from openpyxl import load_workbook
import time
from ....models.brief_case.directories.bank_mfo import bank_mfo
from ....models.brief_case.directories.client_region import client_region
from ....models.brief_case.directories.client_district import client_district
from ....models.brief_case.directories.local_code import local_code
from ....models.brief_case.directories.dis_reg_post_codes import post_codes
from ....models.brief_case.directories.currency import currency
from ....models.brief_case.directories.gender import gender
from ....models.brief_case.directories.credit_satus import status
from ....models.brief_case.directories.directory_interface import DirectoryInterface


def create_directory(Directory, name, db_session):
    
    book = load_workbook(f'project_files/directories/{name}.xlsx',  read_only = True, keep_vba = False)
    sheet = book.active
    rows = sheet.rows
    start = False
    for i in range(0, sheet.max_row):
            
        strokes = next(rows)
        if strokes[0].value == 'code':
            
            start = True
            continue
        if start:
            
            if strokes[0].value != None:
                
                if name=='client_district':
                    get_region = db_session.query(client_region).filter(client_region.post_code == strokes[2].value).first()
                    
                    directory = Directory(post_code = strokes[0].value,
                                        name = strokes[1].value,
                                        code=strokes[2].value,
                                        region_id=get_region and get_region.id
                                )
                elif name=='client_region':
                    directory = Directory(code=strokes[0].value,
                                        post_code = strokes[1].value,
                                        name = strokes[2].value
                    )
                elif name=='bank_mfo':
                    directory = Directory(code=strokes[0].value,
                                        name = strokes[1].value,
                                        region_id = strokes[2].value
                    )
                else:
                    directory = Directory(code=strokes[0].value,
                                        name = strokes[1].value,
                    )
                db_session.add(directory)
    try:
        db_session.commit()
    except:
        db_session.rollback()
        raise
    return {"result":"OK"}



def load_directories(db_session):
    print('start')
    create_directory(client_region, 'client_region', db_session)
    print('client_region')
    create_directory(bank_mfo, 'bank_mfo', db_session)
    print('bank_mfo')
    create_directory(client_district, 'client_district', db_session)
    print('client_district')
    create_directory(currency, 'currency', db_session)
    print('currency')
    create_directory(gender, 'gender', db_session)
    print('gender')
    







def get_directories(db_session):
    return {'bank_mfos': get_directory(bank_mfo, db_session),
            'local_code': get_directory(local_code, db_session),
        'client_region': get_directory(client_region, db_session),
    'currencies' :get_directory(currency, db_session),
    'genders' :get_directory(gender, db_session),
    'loan_status' :get_directory(status, db_session)}
    




def get_directory(object:DirectoryInterface, db_session):
    directory = db_session.query(object).all()
    direct ={}
    for dir in directory:
        direct[dir.code] = dir.id
        
    return direct



def get_client_region(db_session):
    directory = db_session.query(client_region).all()
    direct =[]
    for dir in directory:
        direct.append({"id": dir.id,
                       "code":dir.code,
                       "name":dir.name})
        
    return direct

def get_bank_mfo_by_region(region_id,db_session):
    directory = db_session.query(bank_mfo)
    if region_id is not None:
        directory = directory.filter(bank_mfo.region_id == region_id)
    
    directory = directory.all()
    direct =[]
    for dir in directory:
        direct.append({"id": dir.id,
                       "code":dir.code,
                       "name":dir.name})
        
    return direct




def load_dist_post_codes(db_session):
    book = load_workbook('project_files/directories/post_codes.xlsx',  read_only = True, keep_vba = False)
    sheet = book.active
    rows = sheet.rows
    start = False
    for i in range(0, sheet.max_row):
            
        strokes = next(rows)
        if strokes[0].value == 'dist_code_post':
            
            start = True
            continue
        if start:
            
            if strokes[0].value != None:
                product = post_codes(dist_code_post = strokes[0].value,
                                     reg_code_post = strokes[2].value,
                                     dist_code_iabs = strokes[4].value,
                                     dist_name_iabs = strokes[5].value)
                db_session.add(product)
    try:
        db_session.commit()
    except:
        db_session.rollback()
        raise
    return {"result":"OK"}